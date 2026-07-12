"""Tests for excel_output.py — direct-capitalization workbook generation.

All fixtures are synthetic and anonymous.  No private PDFs or PII are used.
"""
from __future__ import annotations

import re
from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from revenue_kun.excel_output import (
    SHEET_EXPENSE,
    SHEET_OER,
    SHEET_RENT_ROLL,
    DirectCapRow,
    _C_CAM,
    _C_OTHER,
    _C_PARKING,
    _C_RENT,
    _C_UTIL,
    _INCOME_COLS,
    _VACANT_NOTE,
    write_direct_cap_workbook,
)


# ---------------------------------------------------------------------------
# Synthetic fixtures
# ---------------------------------------------------------------------------

def _make_rows() -> list[DirectCapRow]:
    """Three synthetic, anonymous rent roll rows (2 occupied, 1 vacant)."""
    return [
        DirectCapRow(
            区画="101",
            ステータス="入居",
            月額賃料=80_000,
            月額共益費=5_000,
            月額水道光熱費=2_000,
            月額駐車場=10_000,
            月額その他収入=0,
        ),
        DirectCapRow(
            区画="102",
            ステータス="入居",
            月額賃料=75_000,
            月額共益費=5_000,
            月額水道光熱費=2_000,
            月額駐車場=None,
            月額その他収入=None,
        ),
        DirectCapRow(
            区画="103",
            ステータス="空室",
            月額賃料=None,
            月額共益費=None,
            月額水道光熱費=None,
            月額駐車場=None,
            月額その他収入=None,
        ),
    ]


@pytest.fixture(scope="module")
def workbook_path(tmp_path_factory) -> Path:
    """Write the workbook once; reuse across all tests in this module."""
    p = tmp_path_factory.mktemp("excel_output") / "test_direct_cap.xlsx"
    write_direct_cap_workbook(p, _make_rows())
    return p


@pytest.fixture(scope="module")
def wb(workbook_path):
    """Load workbook without data_only so formula strings are accessible."""
    return load_workbook(workbook_path)


# ---------------------------------------------------------------------------
# Helper
# ---------------------------------------------------------------------------

def _find_row_by_label(ws, search_col: int, label: str) -> int | None:
    """Return the first row number where ws.cell(r, search_col).value == label."""
    for row in ws.iter_rows():
        cell = row[search_col - 1]
        if cell.value == label:
            return cell.row
    return None


def _all_formula_strings(ws) -> list[str]:
    """Collect every cell value that is a formula string (starts with '=')."""
    result = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                result.append(cell.value)
    return result


# ---------------------------------------------------------------------------
# Test 1: all three required sheets present
# ---------------------------------------------------------------------------

def test_workbook_has_all_three_sheets(wb):
    assert SHEET_OER in wb.sheetnames, f"{SHEET_OER!r} not in sheetnames"
    assert SHEET_EXPENSE in wb.sheetnames, f"{SHEET_EXPENSE!r} not in sheetnames"
    assert SHEET_RENT_ROLL in wb.sheetnames, f"{SHEET_RENT_ROLL!r} not in sheetnames"


def test_sheet_order(wb):
    """OER must be sheet 0, EXPENSE sheet 1, RENT_ROLL sheet 2."""
    names = wb.sheetnames
    assert names[0] == SHEET_OER
    assert names[1] == SHEET_EXPENSE
    assert names[2] == SHEET_RENT_ROLL


# ---------------------------------------------------------------------------
# Test 2 & 3: monthly and annual total rows exist in 読み取りレントロール
# ---------------------------------------------------------------------------

def test_rent_roll_has_monthly_total_row(wb):
    ws = wb[SHEET_RENT_ROLL]
    monthly_row = _find_row_by_label(ws, 1, "月計")
    assert monthly_row is not None, "月計 row not found in column A"
    # At least one income cell in that row contains a SUM formula
    has_sum = any(
        isinstance(ws.cell(monthly_row, c).value, str)
        and ws.cell(monthly_row, c).value.upper().startswith("=SUM(")
        for c in _INCOME_COLS
    )
    assert has_sum, "Monthly total row has no SUM formula in income columns"


def test_rent_roll_has_annual_total_row(wb):
    ws = wb[SHEET_RENT_ROLL]
    annual_row = _find_row_by_label(ws, 1, "年計")
    assert annual_row is not None, "年計 row not found in column A"


def test_annual_total_row_directly_below_monthly_total_row(wb):
    ws = wb[SHEET_RENT_ROLL]
    monthly_row = _find_row_by_label(ws, 1, "月計")
    annual_row = _find_row_by_label(ws, 1, "年計")
    assert monthly_row is not None and annual_row is not None
    assert annual_row == monthly_row + 1, (
        f"年計 row ({annual_row}) should be directly below 月計 row ({monthly_row})"
    )


# ---------------------------------------------------------------------------
# Test 4: annual total row formulas convert monthly totals to annual (* 12)
# ---------------------------------------------------------------------------

def test_annual_total_row_formulas_use_times_12(wb):
    ws = wb[SHEET_RENT_ROLL]
    annual_row = _find_row_by_label(ws, 1, "年計")
    assert annual_row is not None
    for c in _INCOME_COLS:
        v = ws.cell(annual_row, c).value
        assert isinstance(v, str) and "*12" in v.replace(" ", ""), (
            f"Annual total cell ({get_column_letter(c)}{annual_row}) "
            f"should contain *12, got: {v!r}"
        )


# ---------------------------------------------------------------------------
# Test 5: OER E5:E9 reference the annual total row in 読み取りレントロール
# ---------------------------------------------------------------------------

def test_oer_core_income_cells_reference_rent_roll_annual_row(wb):
    """E5 (賃料) / E6 (共益費) は常に 読み取りレントロール annual row を参照する。"""
    rr_ws = wb[SHEET_RENT_ROLL]
    annual_row = _find_row_by_label(rr_ws, 1, "年計")
    assert annual_row is not None

    oer_ws = wb[SHEET_OER]
    for cell_ref in ("E5", "E6"):   # 賃料・共益費は常に cross-sheet ref
        formula = oer_ws[cell_ref].value
        assert isinstance(formula, str) and formula.startswith("="), (
            f"OER {cell_ref} should be a formula, got: {formula!r}"
        )
        assert SHEET_RENT_ROLL in formula, (
            f"OER {cell_ref} should reference {SHEET_RENT_ROLL!r}, got: {formula!r}"
        )
        assert str(annual_row) in formula, (
            f"OER {cell_ref} should reference row {annual_row}, got: {formula!r}"
        )


def test_oer_optional_income_cells_reference_rent_roll(wb):
    """v0.5.2: 付帯収入（水道代/駐車場/その他）も常にクロスシート参照。opt-in/out 廃止。"""
    rr_ws = wb[SHEET_RENT_ROLL]
    annual_row = _find_row_by_label(rr_ws, 1, "年計")
    assert annual_row is not None

    oer_ws = wb[SHEET_OER]
    for cell_ref in ("E7", "E8", "E9"):
        formula = oer_ws[cell_ref].value
        assert isinstance(formula, str) and formula.startswith("="), (
            f"OER {cell_ref} should be a formula, got: {formula!r}"
        )
        assert formula != "=0", f"OER {cell_ref} must not be =0 (v0.5.2 auto-includes all income)"
        assert SHEET_RENT_ROLL in formula, (
            f"OER {cell_ref} should reference {SHEET_RENT_ROLL!r}, got: {formula!r}"
        )
        assert str(annual_row) in formula, (
            f"OER {cell_ref} should reference row {annual_row}, got: {formula!r}"
        )


def test_oer_optional_income_labels_have_no_exclusion_marker(wb):
    """v0.5.2: 付帯収入行ラベルに「算入対象外」は付与しない（常時算入のため）。"""
    oer_ws = wb[SHEET_OER]
    for row_num in (7, 8, 9):
        label = oer_ws.cell(row_num, 4).value or ""
        assert "算入対象外" not in label, (
            f"OER row {row_num} D column should not contain '算入対象外', got: {label!r}"
        )


def test_oer_optional_income_rent_roll_values_visible_regardless(workbook_path):
    """読み取りレントロールには水道代収入値が常に表示される（opt-out でも）。
    _make_rows() は月額水道光熱費=2_000 を含む — opt-out でも消えてはいけない。"""
    from openpyxl import load_workbook
    wb = load_workbook(workbook_path)
    rr_ws = wb[SHEET_RENT_ROLL]
    # 行2 (101号室) は月額水道光熱費=2_000 が設定されている
    water_col = _C_UTIL  # 5
    assert rr_ws.cell(2, water_col).value == 2_000, (
        "読み取りレントロール row2 水道代収入列は opt-out でも表示されるべき"
    )


# ---------------------------------------------------------------------------
# Test 6: OER E5:E9 formulas do NOT contain *12
# ---------------------------------------------------------------------------

def test_oer_core_income_cells_do_not_multiply_by_12(wb):
    """E5:E9 (賃料/共益費/水道代/駐車場/その他) は annual total 参照なので *12 不要。"""
    oer_ws = wb[SHEET_OER]
    for cell_ref in ("E5", "E6", "E7", "E8", "E9"):
        formula = oer_ws[cell_ref].value or ""
        normalized = formula.replace(" ", "")
        assert "*12" not in normalized, (
            f"OER {cell_ref} must not contain *12 (annual total already in "
            f"{SHEET_RENT_ROLL}), got: {formula!r}"
        )


# ---------------------------------------------------------------------------
# Test 11: OER calculation formulas (E10, E20-E24) — self-computing model
# ---------------------------------------------------------------------------

def test_oer_gpi_total_formula(wb):
    oer_ws = wb[SHEET_OER]
    assert oer_ws["E10"].value == "=SUM(E5:E9)", (
        f"E10 should be =SUM(E5:E9), got: {oer_ws['E10'].value!r}"
    )


def test_oer_egi_formula(wb):
    oer_ws = wb[SHEET_OER]
    assert oer_ws["E20"].value == "=E10*(1-N(E13)-N(E14))", (
        f"E20 (EGI) formula wrong: {oer_ws['E20'].value!r}"
    )


def test_oer_opex_formula_is_egi_times_expense_ratio(wb):
    """Regression guard: operating expenses must use EGI × expense ratio (E20*N(E15))."""
    oer_ws = wb[SHEET_OER]
    assert oer_ws["E21"].value == "=E20*N(E15)", (
        f"E21 (opex) must be =E20*N(E15) (EGI×expense-ratio method), "
        f"got: {oer_ws['E21'].value!r}"
    )


def test_oer_noi_formula(wb):
    oer_ws = wb[SHEET_OER]
    assert oer_ws["E22"].value == "=E20-E21", (
        f"E22 (NOI) formula wrong: {oer_ws['E22'].value!r}"
    )


def test_oer_net_income_formula(wb):
    oer_ws = wb[SHEET_OER]
    assert oer_ws["E23"].value == "=E22-N(E16)", (
        f"E23 (net income) formula wrong: {oer_ws['E23'].value!r}"
    )


def test_oer_indicated_value_formula(wb):
    oer_ws = wb[SHEET_OER]
    assert oer_ws["E24"].value == '=IFERROR(E23/E17,"")', (
        f"E24 (収益試算値) formula wrong: {oer_ws['E24'].value!r}"
    )


# ---------------------------------------------------------------------------
# Test 12: v0.5.2 sheet independence — neither calculation sheet references
# the other by name, anywhere.
# ---------------------------------------------------------------------------

def test_noi_does_not_reference_expense_sheet(wb):
    """費用詳細版 must not be in E22 (NOI)."""
    oer_ws = wb[SHEET_OER]
    noi_formula = oer_ws["E22"].value or ""
    assert SHEET_EXPENSE not in noi_formula, (
        f"E22 (NOI) must not reference {SHEET_EXPENSE!r}. Got: {noi_formula!r}"
    )


def test_oer_sheet_never_references_expense_sheet(wb):
    """v0.5.2: 直接還元法_OER のいかなる数式にも 直接還元法‗費用詳細版 は現れない。"""
    oer_ws = wb[SHEET_OER]
    for formula in _all_formula_strings(oer_ws):
        assert SHEET_EXPENSE not in formula, (
            f"OER sheet formula unexpectedly references {SHEET_EXPENSE!r}: {formula!r}"
        )


def test_expense_sheet_never_references_oer_sheet(wb):
    """v0.5.2: 直接還元法‗費用詳細版 のいかなる数式にも 直接還元法_OER は現れない。"""
    exp_ws = wb[SHEET_EXPENSE]
    for formula in _all_formula_strings(exp_ws):
        assert SHEET_OER not in formula, (
            f"Expense sheet formula unexpectedly references {SHEET_OER!r}: {formula!r}"
        )


def test_both_sheets_only_reference_rent_roll(wb):
    """両計算シートが参照してよい他シートは 読み取りレントロール だけである。"""
    for sheet_name in (SHEET_OER, SHEET_EXPENSE):
        ws = wb[sheet_name]
        for formula in _all_formula_strings(ws):
            if "!" in formula:  # cross-sheet reference syntax
                assert SHEET_RENT_ROLL in formula, (
                    f"{sheet_name} formula references an unexpected sheet: {formula!r}"
                )


# ---------------------------------------------------------------------------
# Test 13: input cells E13:E17 are empty but have border and fill
# ---------------------------------------------------------------------------

def test_oer_input_cells_are_empty(wb):
    oer_ws = wb[SHEET_OER]
    for ref in ("E13", "E14", "E15", "E16", "E17"):
        v = oer_ws[ref].value
        assert v is None, f"OER {ref} should be empty (user input cell), got: {v!r}"


def test_oer_input_cells_have_border(wb):
    oer_ws = wb[SHEET_OER]
    for ref in ("E13", "E14", "E15", "E16", "E17"):
        b = oer_ws[ref].border
        has_border = any(
            s is not None and s != "none"
            for s in (b.left.style, b.right.style, b.top.style, b.bottom.style)
        )
        assert has_border, f"OER {ref} (input cell) should have a border"


def test_oer_input_cells_have_fill(wb):
    oer_ws = wb[SHEET_OER]
    for ref in ("E13", "E14", "E15", "E16", "E17"):
        fill = oer_ws[ref].fill
        assert fill is not None and fill.fgColor is not None, (
            f"OER {ref} (input cell) should have a fill colour"
        )


# ---------------------------------------------------------------------------
# Test 14: 費用詳細版 v0.5.2 layout — independent income block, assumption
# inputs, expense line items, and calculation results (own row numbers,
# no reference to 直接還元法_OER).
# ---------------------------------------------------------------------------

def test_expense_income_block_references_rent_roll(wb):
    """費用詳細版 E5:E9 は 読み取りレントロール を直接参照する（OER 経由ではない）。"""
    rr_ws = wb[SHEET_RENT_ROLL]
    annual_row = _find_row_by_label(rr_ws, 1, "年計")
    assert annual_row is not None

    exp_ws = wb[SHEET_EXPENSE]
    for cell_ref in ("E5", "E6", "E7", "E8", "E9"):
        formula = exp_ws[cell_ref].value
        assert isinstance(formula, str) and formula.startswith("="), (
            f"Expense {cell_ref} should be a formula, got: {formula!r}"
        )
        assert SHEET_RENT_ROLL in formula, (
            f"Expense {cell_ref} should reference {SHEET_RENT_ROLL!r}, got: {formula!r}"
        )
        assert str(annual_row) in formula


def test_expense_gpi_total_formula(wb):
    exp_ws = wb[SHEET_EXPENSE]
    assert exp_ws["E10"].value == "=SUM(E5:E9)", (
        f"費用詳細版 E10 should be =SUM(E5:E9), got: {exp_ws['E10'].value!r}"
    )


def test_expense_assumption_input_cells_are_empty(wb):
    """空室損失率(E13)/貸倒損失率(E14)/資本的支出(E15)/還元利回り(E16)。"""
    exp_ws = wb[SHEET_EXPENSE]
    for ref in ("E13", "E14", "E15", "E16"):
        v = exp_ws[ref].value
        assert v is None, f"費用詳細版 {ref} should be empty (input cell), got: {v!r}"


def test_expense_line_item_input_cells_are_empty(wb):
    """管理費/修繕費/損害保険料/固定資産税/水道光熱費/その他運営費用 (E19:E24)。"""
    exp_ws = wb[SHEET_EXPENSE]
    for ref in ("E19", "E20", "E21", "E22", "E23", "E24"):
        v = exp_ws[ref].value
        assert v is None, f"費用詳細版 {ref} should be empty (input cell), got: {v!r}"


def test_expense_input_cells_have_border_and_fill(wb):
    exp_ws = wb[SHEET_EXPENSE]
    for ref in ("E13", "E14", "E15", "E16", "E19", "E20", "E21", "E22", "E23", "E24"):
        cell = exp_ws[ref]
        b = cell.border
        has_border = any(
            s is not None and s != "none"
            for s in (b.left.style, b.right.style, b.top.style, b.bottom.style)
        )
        assert has_border, f"費用詳細版 {ref} (input cell) should have a border"
        assert cell.fill is not None and cell.fill.fgColor is not None, (
            f"費用詳細版 {ref} (input cell) should have a fill colour"
        )


def test_expense_opex_total_formula(wb):
    exp_ws = wb[SHEET_EXPENSE]
    assert exp_ws["E25"].value == "=SUM(E19:E24)", (
        f"費用詳細版 E25 (運営費用合計) should be =SUM(E19:E24), got: {exp_ws['E25'].value!r}"
    )


def test_expense_egi_formula(wb):
    exp_ws = wb[SHEET_EXPENSE]
    assert exp_ws["E28"].value == "=E10*(1-N(E13)-N(E14))", (
        f"費用詳細版 E28 (EGI) formula wrong: {exp_ws['E28'].value!r}"
    )


def test_expense_noi_formula(wb):
    exp_ws = wb[SHEET_EXPENSE]
    assert exp_ws["E29"].value == "=E28-E25", (
        f"費用詳細版 E29 (NOI) formula wrong: {exp_ws['E29'].value!r}"
    )


def test_expense_net_income_formula(wb):
    exp_ws = wb[SHEET_EXPENSE]
    assert exp_ws["E30"].value == "=E29-N(E15)", (
        f"費用詳細版 E30 (純収益) formula wrong: {exp_ws['E30'].value!r}"
    )


def test_expense_indicated_value_formula(wb):
    exp_ws = wb[SHEET_EXPENSE]
    assert exp_ws["E31"].value == '=IFERROR(E30/E16,"")', (
        f"費用詳細版 E31 (収益試算値) formula wrong: {exp_ws['E31'].value!r}"
    )


# ---------------------------------------------------------------------------
# Test 7: vacant unit 備考 normalized
# ---------------------------------------------------------------------------

def test_vacant_unit_biko_is_normalized(wb):
    ws = wb[SHEET_RENT_ROLL]
    from revenue_kun.excel_output import _C_STATUS, _C_NOTES
    vacant_notes = []
    for row in ws.iter_rows(min_row=2):
        status = row[_C_STATUS - 1].value
        note = row[_C_NOTES - 1].value
        if status == "空室":
            vacant_notes.append(note)
    assert vacant_notes, "No vacant rows found in test fixture"
    for note in vacant_notes:
        assert note == _VACANT_NOTE, (
            f"Vacant unit 備考 should be {_VACANT_NOTE!r}, got {note!r}"
        )


# ---------------------------------------------------------------------------
# Test 8: money columns use thousands-separator number format
# ---------------------------------------------------------------------------

def test_income_columns_have_thousands_separator_format(wb):
    ws = wb[SHEET_RENT_ROLL]
    from revenue_kun.excel_output import _MONEY_FORMAT
    # Check a data row (row 2) and the monthly total row
    monthly_row = _find_row_by_label(ws, 1, "月計")
    rows_to_check = [2, monthly_row]
    for r in rows_to_check:
        if r is None:
            continue
        for c in _INCOME_COLS:
            fmt = ws.cell(r, c).number_format
            assert _MONEY_FORMAT in fmt, (
                f"Cell {get_column_letter(c)}{r} number_format should contain "
                f"{_MONEY_FORMAT!r}, got {fmt!r}"
            )


# ---------------------------------------------------------------------------
# Test 9: no obvious formula errors in any sheet
# ---------------------------------------------------------------------------

_ERROR_PATTERNS = re.compile(r"#(REF|VALUE|NAME|DIV/0|N/A|NULL|NUM)!", re.IGNORECASE)


def test_no_formula_errors_in_workbook(wb):
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    assert not _ERROR_PATTERNS.search(cell.value), (
                        f"Formula error in {sheet_name}!{cell.coordinate}: {cell.value!r}"
                    )


# ---------------------------------------------------------------------------
# Test 10: annual total row cells have borders (no unnecessary borders below)
# ---------------------------------------------------------------------------

def test_annual_total_row_has_borders(wb):
    ws = wb[SHEET_RENT_ROLL]
    annual_row = _find_row_by_label(ws, 1, "年計")
    assert annual_row is not None
    for c in _INCOME_COLS:
        cell = ws.cell(annual_row, c)
        b = cell.border
        assert b.left.style is not None or b.right.style is not None \
            or b.top.style is not None or b.bottom.style is not None, (
            f"Annual total cell {get_column_letter(c)}{annual_row} should have a border"
        )


def test_no_borders_below_annual_total_row(wb):
    ws = wb[SHEET_RENT_ROLL]
    annual_row = _find_row_by_label(ws, 1, "年計")
    assert annual_row is not None
    check_row = annual_row + 1
    for c in _INCOME_COLS:
        cell = ws.cell(check_row, c)
        b = cell.border
        has_border = any(
            s is not None and s != "none"
            for s in (b.left.style, b.right.style, b.top.style, b.bottom.style)
        )
        assert not has_border, (
            f"Row {check_row} col {get_column_letter(c)} should have no border "
            f"(it is below the annual total row)"
        )


# ---------------------------------------------------------------------------
# DirectCapRow.from_rent_roll_unit converter
# ---------------------------------------------------------------------------

def test_from_rent_roll_unit_occupied():
    from revenue_kun.rent_roll import RentRollUnit
    unit = RentRollUnit(
        区画="201",
        用途="住居",
        賃借人=None,
        専有面積_m2=30.0,
        月額賃料_円=60_000,
        月額共益費_円=3_000,
        稼働状況="入居",
        契約満了日=None,
    )
    row = DirectCapRow.from_rent_roll_unit(unit)
    assert row.区画 == "201"
    assert row.ステータス == "入居"
    assert row.月額賃料 == 60_000
    assert row.月額共益費 == 3_000
    assert row.月額水道光熱費 is None
    assert row.備考 is None  # occupied — no note


def test_from_rent_roll_unit_vacant():
    from revenue_kun.rent_roll import RentRollUnit
    unit = RentRollUnit(
        区画="202",
        用途=None,
        賃借人=None,
        専有面積_m2=None,
        月額賃料_円=None,
        月額共益費_円=None,
        稼働状況="空室",
        契約満了日=None,
    )
    row = DirectCapRow.from_rent_roll_unit(unit)
    assert row.ステータス == "空室"
    assert row.月額賃料 is None
    assert row.備考 == _VACANT_NOTE
