"""Tests for POST /api/generate (Issue #82).

Uses only synthetic fixtures already committed to the repo
(`data/dummy_rent_roll.csv`, `data/sample_rentroll_simple.pdf`) or
generated on the fly via `revenue_kun.sample_pdf`. No real property or
tenant information is used.
"""
from __future__ import annotations

import io
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from revenue_kun.excel_output import SHEET_EXPENSE, SHEET_OER, SHEET_RENT_ROLL
from revenue_kun.sample_pdf import build_text_only_pdf
from webui.app import app

_REPO_ROOT = Path(__file__).resolve().parents[1]
_DUMMY_CSV = _REPO_ROOT / "data" / "dummy_rent_roll.csv"
_SIMPLE_PDF = _REPO_ROOT / "data" / "sample_rentroll_simple.pdf"

_XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest.fixture()
def client() -> TestClient:
    return TestClient(app)


def _generate(client: TestClient, filename: str, content: bytes, content_type: str, optional_income=None):
    data = {}
    if optional_income:
        data["optional_income"] = optional_income
    return client.post(
        "/api/generate",
        files={"file": (filename, content, content_type)},
        data=data,
    )


def _load_workbook_from_response(response) -> "load_workbook":
    return load_workbook(io.BytesIO(response.content))


# ---------------------------------------------------------------------------
# Success: CSV / PDF workbook download
# ---------------------------------------------------------------------------

def test_valid_csv_generate_returns_workbook(client):
    response = _generate(client, "rentroll.csv", _DUMMY_CSV.read_bytes(), "text/csv")
    assert response.status_code == 200
    assert response.headers["content-type"] == _XLSX_MEDIA_TYPE


def test_valid_pdf_generate_returns_workbook(client):
    response = _generate(client, "rentroll.pdf", _SIMPLE_PDF.read_bytes(), "application/pdf")
    assert response.status_code == 200
    assert response.headers["content-type"] == _XLSX_MEDIA_TYPE


def test_generate_response_filename_is_direct_cap_xlsx(client):
    response = _generate(client, "rentroll.csv", _DUMMY_CSV.read_bytes(), "text/csv")
    assert 'filename="direct_cap.xlsx"' in response.headers["content-disposition"]


def test_generate_workbook_has_three_expected_sheets(client):
    response = _generate(client, "rentroll.csv", _DUMMY_CSV.read_bytes(), "text/csv")
    wb = _load_workbook_from_response(response)
    assert wb.sheetnames == [SHEET_OER, SHEET_EXPENSE, SHEET_RENT_ROLL]


# ---------------------------------------------------------------------------
# Optional-income -> OptionalIncomeConfig mapping
# ---------------------------------------------------------------------------

def _oer_income_formula(wb, cell: str) -> str:
    return wb[SHEET_OER][cell].value


def test_no_optional_income_selection_excludes_all_from_gpi(client):
    response = _generate(client, "rentroll.csv", _DUMMY_CSV.read_bytes(), "text/csv")
    wb = _load_workbook_from_response(response)
    assert _oer_income_formula(wb, "E7") == "=0"
    assert _oer_income_formula(wb, "E8") == "=0"
    assert _oer_income_formula(wb, "E9") == "=0"


def test_water_income_opt_in_links_to_rent_roll(client):
    response = _generate(client, "rentroll.csv", _DUMMY_CSV.read_bytes(), "text/csv", ["water_income"])
    wb = _load_workbook_from_response(response)
    assert _oer_income_formula(wb, "E7") != "=0"
    assert SHEET_RENT_ROLL in str(_oer_income_formula(wb, "E7"))
    assert _oer_income_formula(wb, "E8") == "=0"
    assert _oer_income_formula(wb, "E9") == "=0"


def test_parking_income_opt_in_links_to_rent_roll(client):
    response = _generate(client, "rentroll.csv", _DUMMY_CSV.read_bytes(), "text/csv", ["parking_income"])
    wb = _load_workbook_from_response(response)
    assert _oer_income_formula(wb, "E7") == "=0"
    assert _oer_income_formula(wb, "E8") != "=0"
    assert _oer_income_formula(wb, "E9") == "=0"


def test_other_income_opt_in_links_to_rent_roll(client):
    response = _generate(client, "rentroll.csv", _DUMMY_CSV.read_bytes(), "text/csv", ["other_income"])
    wb = _load_workbook_from_response(response)
    assert _oer_income_formula(wb, "E7") == "=0"
    assert _oer_income_formula(wb, "E8") == "=0"
    assert _oer_income_formula(wb, "E9") != "=0"


def test_multiple_optional_income_selection(client):
    response = _generate(
        client, "rentroll.csv", _DUMMY_CSV.read_bytes(), "text/csv",
        ["water_income", "parking_income", "other_income"],
    )
    wb = _load_workbook_from_response(response)
    assert _oer_income_formula(wb, "E7") != "=0"
    assert _oer_income_formula(wb, "E8") != "=0"
    assert _oer_income_formula(wb, "E9") != "=0"


def test_invalid_optional_income_category_is_safe_failure(client):
    response = _generate(client, "rentroll.csv", _DUMMY_CSV.read_bytes(), "text/csv", ["not_a_real_category"])
    assert response.status_code == 400
    body = response.json()
    assert body["ok"] is False
    assert body["error"]["detail_code"] == "invalid_optional_income_category"


# ---------------------------------------------------------------------------
# Validation / safe failure (mirrors /api/preview's contract)
# ---------------------------------------------------------------------------

def test_unsupported_extension_is_safe_failure(client):
    response = _generate(client, "rentroll.xlsx", b"whatever", "application/octet-stream")
    assert response.status_code == 400
    body = response.json()
    assert body["ok"] is False
    assert "Traceback" not in response.text


def test_no_file_attached_is_safe_failure(client):
    response = client.post("/api/generate")
    assert response.status_code == 400
    body = response.json()
    assert body["ok"] is False


def test_malformed_pdf_signature_is_safe_failure(client):
    response = _generate(client, "fake.pdf", b"this is not a pdf file at all", "application/pdf")
    assert response.status_code == 400
    body = response.json()
    assert body["ok"] is False
    assert body["error"]["detail_code"] == "invalid_pdf_signature"


def test_scanned_or_table_less_pdf_is_safe_failure(client, tmp_path):
    pdf_path = tmp_path / "no_table.pdf"
    build_text_only_pdf(pdf_path)
    response = _generate(client, "scanned.pdf", pdf_path.read_bytes(), "application/pdf")
    assert response.status_code == 422
    body = response.json()
    assert body["ok"] is False
    assert body["error"]["detail_code"] == "rent_roll_table_not_found"


def test_oversized_upload_is_safe_failure(client, monkeypatch):
    monkeypatch.setenv("REVENUE_KUN_MAX_UPLOAD_MB", "1")
    oversized = b"x" * (2 * 1024 * 1024)
    response = _generate(client, "big.csv", oversized, "text/csv")
    assert response.status_code == 400
    body = response.json()
    assert body["ok"] is False
    assert body["error"]["detail_code"] == "upload_too_large"


def test_safe_failure_never_returns_workbook_content_type(client):
    response = _generate(client, "rentroll.xlsx", b"whatever", "application/octet-stream")
    assert response.headers["content-type"] != _XLSX_MEDIA_TYPE


def test_safe_failure_response_has_no_filesystem_path(client):
    response = _generate(client, "rentroll.xlsx", b"whatever", "application/octet-stream")
    assert "revenue_kun_webui_" not in response.text
    assert "Traceback" not in response.text


# ---------------------------------------------------------------------------
# Temporary-file cleanup
# ---------------------------------------------------------------------------

def _temp_dir_count() -> int:
    import tempfile

    base = Path(tempfile.gettempdir())
    return len(list(base.glob("revenue_kun_webui_*")))


def test_temp_files_removed_after_successful_generate(client):
    before = _temp_dir_count()
    _generate(client, "rentroll.csv", _DUMMY_CSV.read_bytes(), "text/csv")
    assert _temp_dir_count() == before


def test_temp_files_removed_after_generate_validation_failure(client):
    before = _temp_dir_count()
    _generate(client, "rentroll.xlsx", b"whatever", "application/octet-stream")
    assert _temp_dir_count() == before


def test_temp_files_removed_after_generate_extraction_failure(client, tmp_path):
    pdf_path = tmp_path / "no_table.pdf"
    build_text_only_pdf(pdf_path)
    before = _temp_dir_count()
    _generate(client, "scanned.pdf", pdf_path.read_bytes(), "application/pdf")
    assert _temp_dir_count() == before


# ---------------------------------------------------------------------------
# Regression: preview endpoint still works
# ---------------------------------------------------------------------------

def test_preview_endpoint_still_works_after_generate_added(client):
    response = client.post(
        "/api/preview",
        files={"file": ("rentroll.csv", _DUMMY_CSV.read_bytes(), "text/csv")},
    )
    assert response.status_code == 200
    assert response.json()["ok"] is True
