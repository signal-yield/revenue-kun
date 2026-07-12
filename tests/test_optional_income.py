"""付帯収入（optional income）機能のテスト。

検証項目:
  A. 後方互換性 — optional_income 設定なし / false / columns: [] で GPI 変わらず
  B. water opt-in — GPI が 8,211,540 円になること
  C. water opt-out — 同入力で GPI が 7,764,000 円になること
  D. カテゴリ分離 — 水道代（収入）と運営費用の水道光熱費（費用）を混同しない
  E. PDF 抽出 — "水道代" ヘッダーが water として認識されること
  F. Excel 出力 — opt-in 時は月額水道光熱費が設定され、opt-out 時は None のまま
"""
from __future__ import annotations

from pathlib import Path

import pytest

from revenue_kun.config import (
    Assumptions,
    OptionalIncomeConfig,
    load_assumptions,
)
from revenue_kun.excel_output import DirectCapRow
from revenue_kun.noi import compute_noi
from revenue_kun.pdf_extract import _resolve_header_key, extract_rent_roll_from_pdf
from revenue_kun.rent_roll import RentRollUnit
from revenue_kun.sample_pdf import build_pdf


# ---------------------------------------------------------------------------
# ヘルパー
# ---------------------------------------------------------------------------

def _make_assumptions(
    include_in_gpi: bool = False,
    columns: list[str] | None = None,
) -> Assumptions:
    """テスト用 Assumptions を生成する（cap_rate などは最小値で設定）。"""
    a = load_assumptions("assumptions.sample.yaml")
    a.optional_income = OptionalIncomeConfig(
        include_in_gpi=include_in_gpi,
        columns=columns or [],
    )
    return a


def _unit(
    room: str = "101",
    rent: float = 601_000,
    cam: float = 46_000,
    water: float | None = None,
    parking: float | None = None,
    other_income: float | None = None,
    status: str = "入居",
) -> RentRollUnit:
    return RentRollUnit(
        区画=room,
        用途="住宅",
        賃借人=None,
        専有面積_m2=30.0,
        月額賃料_円=rent,
        月額共益費_円=cam,
        稼働状況=status,
        契約満了日=None,
        月額水道代_円=water,
        月額駐車場収入_円=parking,
        月額その他収入_円=other_income,
    )


# ---------------------------------------------------------------------------
# A. 後方互換性テスト
# ---------------------------------------------------------------------------

class TestBackwardCompatibility:
    """optional_income 設定がない / false / empty でも既存 GPI が変わらない。"""

    def test_no_section_defaults_to_false(self):
        """assumptions に optional_income セクションがなければ include_in_gpi=False。"""
        a = load_assumptions("assumptions.sample.yaml")
        # sample.yaml には optional_income: include_in_gpi: false が入っている
        assert a.optional_income.include_in_gpi is False
        assert a.optional_income.columns == []

    def test_include_false_gpi_unchanged(self):
        """include_in_gpi=False のとき water があっても GPI に算入しない。"""
        a = _make_assumptions(include_in_gpi=False, columns=["water"])
        units = [_unit(rent=601_000, cam=46_000, water=37_295)]
        noi = compute_noi(units, a)
        expected_gpi = (601_000 + 46_000) * 12
        assert noi.gpi == pytest.approx(expected_gpi)

    def test_empty_columns_gpi_unchanged(self):
        """include_in_gpi=True でも columns=[] なら付帯収入を算入しない。"""
        a = _make_assumptions(include_in_gpi=True, columns=[])
        units = [_unit(rent=601_000, cam=46_000, water=37_295)]
        noi = compute_noi(units, a)
        expected_gpi = (601_000 + 46_000) * 12
        assert noi.gpi == pytest.approx(expected_gpi)

    def test_rent_income_and_cam_income_breakdown(self):
        """収入内訳フィールドが正しく設定される（opt-out）。"""
        a = _make_assumptions(include_in_gpi=False)
        units = [_unit(rent=601_000, cam=46_000, water=37_295)]
        noi = compute_noi(units, a)
        assert noi.rent_income == pytest.approx(601_000 * 12)
        assert noi.cam_income == pytest.approx(46_000 * 12)
        assert noi.optional_income_total == pytest.approx(0.0)


# ---------------------------------------------------------------------------
# B. water opt-in テスト
# ---------------------------------------------------------------------------

class TestWaterOptin:
    """水道代を opt-in すると GPI = 8,211,540円になること。"""

    def test_gpi_with_water_optin(self):
        """グリーン蛍 PDF 相当: rent+cam+water × 17区画 ≈ 8,211,540。"""
        a = _make_assumptions(include_in_gpi=True, columns=["water"])
        units = [_unit(rent=601_000, cam=46_000, water=37_295)]
        noi = compute_noi(units, a)
        # 1区画の場合: (601000 + 46000 + 37295) * 12 = 8,211,540
        expected = (601_000 + 46_000 + 37_295) * 12
        assert noi.gpi == pytest.approx(expected)

    def test_income_breakdown_with_water_optin(self):
        """opt-in 時は optional_income_total が正しく設定される。"""
        a = _make_assumptions(include_in_gpi=True, columns=["water"])
        units = [_unit(rent=601_000, cam=46_000, water=37_295)]
        noi = compute_noi(units, a)
        assert noi.rent_income == pytest.approx(601_000 * 12)
        assert noi.cam_income == pytest.approx(46_000 * 12)
        assert noi.optional_income_total == pytest.approx(37_295 * 12)
        assert noi.gpi == pytest.approx(noi.rent_income + noi.cam_income + noi.optional_income_total)

    def test_multi_unit_water_optin(self):
        """複数区画の場合も水道代が正しく合計される。"""
        a = _make_assumptions(include_in_gpi=True, columns=["water"])
        units = [
            _unit("101", rent=100_000, cam=10_000, water=5_000),
            _unit("102", rent=120_000, cam=12_000, water=6_000),
            _unit("103", rent=80_000, cam=8_000, water=None),  # 水道代なし
        ]
        noi = compute_noi(units, a)
        expected_gpi = (100_000 + 10_000 + 5_000 + 120_000 + 12_000 + 6_000 + 80_000 + 8_000) * 12
        assert noi.gpi == pytest.approx(expected_gpi)
        assert noi.optional_income_total == pytest.approx((5_000 + 6_000) * 12)

    def test_water_optin_note_added(self):
        """opt-in 時は notes に付帯収入算入の記録が入る。"""
        a = _make_assumptions(include_in_gpi=True, columns=["water"])
        units = [_unit(water=37_295)]
        noi = compute_noi(units, a)
        assert any("opt-in" in n for n in noi.notes)


# ---------------------------------------------------------------------------
# C. water opt-out テスト
# ---------------------------------------------------------------------------

class TestWaterOptout:
    """同入力で include_in_gpi=false → GPI = 7,764,000 円になること。"""

    def test_gpi_without_water_optout(self):
        a = _make_assumptions(include_in_gpi=False)
        units = [_unit(rent=601_000, cam=46_000, water=37_295)]
        noi = compute_noi(units, a)
        expected = (601_000 + 46_000) * 12
        assert noi.gpi == pytest.approx(expected)
        assert noi.optional_income_total == pytest.approx(0.0)

    def test_water_not_in_columns_is_excluded(self):
        """include_in_gpi=True でも water が columns にない場合は除外される。"""
        a = _make_assumptions(include_in_gpi=True, columns=["parking"])
        units = [_unit(rent=601_000, cam=46_000, water=37_295)]
        noi = compute_noi(units, a)
        expected = (601_000 + 46_000) * 12  # water は除外
        assert noi.gpi == pytest.approx(expected)


# ---------------------------------------------------------------------------
# D. カテゴリ分離テスト
# ---------------------------------------------------------------------------

class TestCategorySeparation:
    """水道代（収入サイド）と運営費用の水道光熱費（費用サイド）を混同しない。"""

    def test_water_income_is_not_opex(self):
        """RentRollUnit.月額水道代_円 は収入フィールド。opex に影響しない。"""
        a = _make_assumptions(include_in_gpi=True, columns=["water"])
        # 運営費用に水道光熱費がある
        a.opex["水道光熱費"] = 1_200_000
        units = [_unit(rent=100_000, cam=10_000, water=5_000)]
        noi = compute_noi(units, a)
        # GPI には水道代（収入）が算入される
        assert noi.optional_income_total == pytest.approx(5_000 * 12)
        # 運営費用には水道光熱費（費用）が算入される（別物）
        assert "水道光熱費" in noi.opex_breakdown
        assert noi.opex_breakdown["水道光熱費"] == 1_200_000

    def test_get_optional_income_method(self):
        """RentRollUnit.get_optional_income() が canonical key で正しく取得できる。"""
        u = _unit(water=37_295, parking=20_000, other_income=5_000)
        assert u.get_optional_income("water") == 37_295
        assert u.get_optional_income("parking") == 20_000
        assert u.get_optional_income("other_income") == 5_000
        assert u.get_optional_income("unknown") is None

    def test_vacant_unit_water_excluded_from_gpi(self):
        """空室区画の水道代は GPI から除外される（稼働区画のみ算入）。"""
        a = _make_assumptions(include_in_gpi=True, columns=["water"])
        units = [
            _unit("101", rent=100_000, cam=10_000, water=5_000, status="入居"),
            _unit("102", rent=None, cam=None, water=5_000, status="空室"),  # 空室
        ]
        noi = compute_noi(units, a)
        # 空室区画の水道代は算入されない
        assert noi.optional_income_total == pytest.approx(5_000 * 12)


# ---------------------------------------------------------------------------
# E. PDF 抽出テスト
# ---------------------------------------------------------------------------

class TestPdfExtraction:
    """水道代ヘッダーが water として認識され、RentRollUnit に保持される。"""

    def test_water_header_resolved(self):
        """「水道代」ヘッダーが canonical key 'water' に解決される。"""
        assert _resolve_header_key("水道代") == "water"

    def test_water_fee_header_resolved(self):
        """「水道費」も water に解決される。"""
        assert _resolve_header_key("水道費") == "water"

    def test_parking_revenue_header_resolved(self):
        """「駐車場収入」が parking に解決される。"""
        assert _resolve_header_key("駐車場収入") == "parking"

    def test_parking_header_resolved(self):
        """「駐車場」が parking に解決される。"""
        assert _resolve_header_key("駐車場") == "parking"

    def test_other_income_header_resolved(self):
        """「その他収入」が other_income に解決される。"""
        assert _resolve_header_key("その他収入") == "other_income"

    def test_existing_headers_unaffected(self):
        """既存ヘッダーのマッピングが壊れていない。"""
        assert _resolve_header_key("部屋番号") == "room"
        assert _resolve_header_key("月額賃料(円)") == "rent"
        assert _resolve_header_key("共益費") == "cam"
        assert _resolve_header_key("入居/空室") == "status"
        assert _resolve_header_key("備考") == "notes"

    def test_water_column_extracted_to_rent_roll_unit(self, tmp_path):
        """水道代列を含む PDF から RentRollUnit.月額水道代_円 が抽出される。"""
        headers = ["部屋番号", "用途", "面積(㎡)", "月額賃料(円)", "共益費(円)", "入居/空室", "水道代"]
        rows = [["101", "住宅", "30.0", "100,000", "10,000", "入居", "5,000"]]
        p = tmp_path / "water_test.pdf"
        build_pdf(p, headers, rows)
        units, rep = extract_rent_roll_from_pdf(p)
        assert len(units) == 1
        assert units[0].月額水道代_円 == 5_000
        assert "water" in rep.optional_income_found

    def test_no_water_column_unit_water_is_none(self, tmp_path):
        """水道代列がない PDF では RentRollUnit.月額水道代_円 が None。"""
        headers = ["部屋番号", "用途", "面積(㎡)", "月額賃料(円)", "共益費(円)", "入居/空室"]
        rows = [["101", "住宅", "30.0", "100,000", "10,000", "入居"]]
        p = tmp_path / "no_water.pdf"
        build_pdf(p, headers, rows)
        units, rep = extract_rent_roll_from_pdf(p)
        assert units[0].月額水道代_円 is None
        assert "water" not in rep.optional_income_found

    def test_optional_income_found_reported(self, tmp_path):
        """水道代・駐車場収入の両列を持つ PDF で optional_income_found が正しく設定される。"""
        headers = [
            "部屋番号", "用途", "面積(㎡)", "月額賃料(円)", "共益費(円)", "入居/空室",
            "水道代", "駐車場収入",
        ]
        rows = [["101", "住宅", "30.0", "100,000", "10,000", "入居", "5,000", "20,000"]]
        p = tmp_path / "multi_oi.pdf"
        build_pdf(p, headers, rows)
        units, rep = extract_rent_roll_from_pdf(p)
        assert "water" in rep.optional_income_found
        assert "parking" in rep.optional_income_found
        assert units[0].月額水道代_円 == 5_000
        assert units[0].月額駐車場収入_円 == 20_000

    def test_existing_extraction_unaffected(self, tmp_path):
        """通常の PDF 抽出（水道代なし）に影響がない（回帰テスト）。"""
        headers = ["部屋番号", "用途", "面積(㎡)", "月額賃料(円)", "共益費(円)", "入居/空室"]
        rows = [
            ["101", "事務所", "50.0", "200,000", "20,000", "入居"],
            ["102", "事務所", "50.0", "200,000", "20,000", "入居"],
        ]
        p = tmp_path / "normal.pdf"
        build_pdf(p, headers, rows)
        units, rep = extract_rent_roll_from_pdf(p)
        assert rep.rows_extracted == 2
        assert all(u.月額賃料_円 == 200_000 for u in units)
        assert rep.optional_income_found == []


# ---------------------------------------------------------------------------
# F. Excel 出力テスト
# ---------------------------------------------------------------------------

class TestExcelOutput:
    """v0.5.2: direct_cap.xlsx の付帯収入は常に両計算シートへ自動算入される。

    設計:
      - from_rent_roll_unit: 抽出値を常に DirectCapRow に格納
      - write_direct_cap_workbook(oi_config=...): oi_config は後方互換のため
        受理するのみで、値には一切影響しない。E7-E9 は常に読み取りレントロール
        へのクロスシート参照になる（opt-in/opt-out という概念自体が廃止された）。
    """

    def _make_unit_with_water(self, water: float | None = 37_295) -> RentRollUnit:
        return RentRollUnit(
            区画="101", 用途="住宅", 賃借人=None, 専有面積_m2=30.0,
            月額賃料_円=601_000, 月額共益費_円=46_000,
            稼働状況="入居", 契約満了日=None,
            月額水道代_円=water,
        )

    # --- from_rent_roll_unit: 常に抽出値を格納 --------------------------------

    def test_from_rent_roll_always_shows_extracted_water(self):
        """水道代がある unit では、opt-in/out 関係なく月額水道光熱費 = 抽出値。"""
        unit = self._make_unit_with_water(water=37_295)
        row = DirectCapRow.from_rent_roll_unit(unit)
        assert row.月額水道光熱費 == 37_295

    def test_from_rent_roll_no_water_in_unit_stays_none(self):
        """水道代が抽出されていない unit では月額水道光熱費 = None のまま。"""
        unit = self._make_unit_with_water(water=None)
        row = DirectCapRow.from_rent_roll_unit(unit)
        assert row.月額水道光熱費 is None

    def test_from_rent_roll_all_optional_income_always_populated(self):
        """水道代・駐車場・その他収入が全て抽出値で格納される。"""
        unit = RentRollUnit(
            区画="103", 用途="住宅", 賃借人=None, 専有面積_m2=30.0,
            月額賃料_円=100_000, 月額共益費_円=10_000,
            稼働状況="入居", 契約満了日=None,
            月額水道代_円=5_000,
            月額駐車場収入_円=20_000,
            月額その他収入_円=3_000,
        )
        row = DirectCapRow.from_rent_roll_unit(unit)
        assert row.月額水道光熱費 == 5_000
        assert row.月額駐車場 == 20_000
        assert row.月額その他収入 == 3_000

    def test_from_rent_roll_parking_always_set(self):
        """駐車場収入も常に格納される。"""
        unit = RentRollUnit(
            区画="102", 用途="住宅", 賃借人=None, 専有面積_m2=30.0,
            月額賃料_円=100_000, 月額共益費_円=10_000,
            稼働状況="入居", 契約満了日=None,
            月額駐車場収入_円=20_000,
        )
        row = DirectCapRow.from_rent_roll_unit(unit)
        assert row.月額駐車場 == 20_000
        assert row.月額水道光熱費 is None  # 水道代は抽出されていない

    # --- workbook レベル: OER formula 制御 ------------------------------------

    def _build_workbook(self, tmp_path, water: float | None, oi_config):
        """水道代を持つ1区画の workbook を生成してパスを返す。"""
        from pathlib import Path
        from revenue_kun.excel_output import write_direct_cap_workbook, DirectCapRow
        unit = self._make_unit_with_water(water=water)
        rows = [DirectCapRow.from_rent_roll_unit(unit)]
        p = Path(tmp_path) / "test.xlsx"
        write_direct_cap_workbook(p, rows, oi_config=oi_config)
        return p

    def _load_wb(self, path):
        from openpyxl import load_workbook
        return load_workbook(path)

    def test_workbook_water_oer_always_references_rent_roll_regardless_of_config(self, tmp_path):
        """v0.5.2: oi_config に関わらず、OER E7（水道代収入）は常にクロスシート参照。"""
        from revenue_kun.excel_output import SHEET_OER, SHEET_RENT_ROLL
        for oi in (
            None,
            OptionalIncomeConfig(include_in_gpi=False),
            OptionalIncomeConfig(include_in_gpi=True, columns=["water"]),
            OptionalIncomeConfig(include_in_gpi=True, columns=["parking"]),
        ):
            p = self._build_workbook(tmp_path, water=37_295, oi_config=oi)
            oer_ws = self._load_wb(p)[SHEET_OER]
            formula = oer_ws["E7"].value or ""
            assert formula != "=0", f"E7 must never be =0 (config={oi!r}), got: {formula!r}"
            assert SHEET_RENT_ROLL in formula

    def test_workbook_rent_roll_shows_water(self, tmp_path):
        """読み取りレントロール には水道代収入が常に表示される。"""
        from revenue_kun.excel_output import SHEET_RENT_ROLL, _C_UTIL
        oi = OptionalIncomeConfig(include_in_gpi=False)
        p = self._build_workbook(tmp_path, water=37_295, oi_config=oi)
        rr_ws = self._load_wb(p)[SHEET_RENT_ROLL]
        assert rr_ws.cell(2, _C_UTIL).value == 37_295

    def test_workbook_label_never_contains_excluded(self, tmp_path):
        """v0.5.2: OER 付帯収入行ラベルに「算入対象外」は付与されない。"""
        from revenue_kun.excel_output import SHEET_OER
        oi = OptionalIncomeConfig(include_in_gpi=False)
        p = self._build_workbook(tmp_path, water=37_295, oi_config=oi)
        oer_ws = self._load_wb(p)[SHEET_OER]
        label = oer_ws.cell(7, 4).value or ""
        assert "算入対象外" not in label, f"D7 should not contain '算入対象外', got: {label!r}"

    def test_workbook_parking_also_always_references_rent_roll(self, tmp_path):
        """water のみを指す旧 columns=["water"] を渡しても、E8（駐車場収入）も常にクロスシート参照。"""
        from revenue_kun.excel_output import SHEET_OER, SHEET_RENT_ROLL
        oi = OptionalIncomeConfig(include_in_gpi=True, columns=["water"])
        p = self._build_workbook(tmp_path, water=37_295, oi_config=oi)
        oer_ws = self._load_wb(p)[SHEET_OER]
        formula = oer_ws["E8"].value or ""
        assert formula != "=0", f"E8 must never be =0, got: {formula!r}"
        assert SHEET_RENT_ROLL in formula

    def test_workbook_gpi_formula_sums_e5_to_e9(self, tmp_path):
        """GPI (E10) は常に =SUM(E5:E9)。"""
        from revenue_kun.excel_output import SHEET_OER
        oi = OptionalIncomeConfig(include_in_gpi=False)
        p = self._build_workbook(tmp_path, water=37_295, oi_config=oi)
        oer_ws = self._load_wb(p)[SHEET_OER]
        assert oer_ws["E10"].value == "=SUM(E5:E9)"

    def test_workbook_none_oi_config_same_gpi_formula_as_legacy_config(self, tmp_path):
        """oi_config 省略・空リスト・全指定のいずれでも同じ数式になる（後方互換）。"""
        from revenue_kun.excel_output import SHEET_OER
        configs = [
            None,
            OptionalIncomeConfig(),
            OptionalIncomeConfig(include_in_gpi=False, columns=[]),
            OptionalIncomeConfig(include_in_gpi=True, columns=["water", "parking", "other_income"]),
        ]
        formulas = []
        for oi in configs:
            p = self._build_workbook(tmp_path, water=37_295, oi_config=oi)
            oer_ws = self._load_wb(p)[SHEET_OER]
            formulas.append(tuple(oer_ws[ref].value for ref in ("E5", "E6", "E7", "E8", "E9", "E10")))
        assert len(set(formulas)) == 1, f"All configs should yield identical formulas, got: {formulas!r}"


# ---------------------------------------------------------------------------
# config 読み込みテスト
# ---------------------------------------------------------------------------

class TestOptionalIncomeConfig:
    """OptionalIncomeConfig の読み込みと検証。"""

    def test_load_defaults_from_yaml(self):
        """sample.yaml から OptionalIncomeConfig がデフォルト値で読み込まれる。"""
        a = load_assumptions("assumptions.sample.yaml")
        assert a.optional_income.include_in_gpi is False
        assert a.optional_income.columns == []

    def test_manual_override(self):
        """OptionalIncomeConfig を直接設定できる。"""
        cfg = OptionalIncomeConfig(include_in_gpi=True, columns=["water", "parking"])
        assert cfg.include_in_gpi is True
        assert "water" in cfg.columns
        assert "parking" in cfg.columns

    def test_default_config_is_false(self):
        """OptionalIncomeConfig() のデフォルトは include_in_gpi=False。"""
        cfg = OptionalIncomeConfig()
        assert cfg.include_in_gpi is False
        assert cfg.columns == []
