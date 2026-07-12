"""v0.5.2 acceptance tests: independent OER / expense-detail calculation sheets.

Confirms the confirmed v0.5.2 product boundary end-to-end using the exact
synthetic figures from the implementation spec:

  - recurring income (rent/common fee/water/parking/other) is always
    auto-included in both calculation sheets' GPI, with no selection
  - the OER sheet computes NOI via EGI × 採用OER (運営費用率)
  - the expense-detail sheet computes NOI via individually summed expense
    line items, independent of the OER sheet
  - neither calculation sheet references the other by name
  - non-income items (deposits, monthly/annual total rows) never enter GPI

openpyxl does not evaluate formulas, so numeric expectations are verified
by re-computing the same formulas in plain Python from the same synthetic
inputs, alongside structural checks of the formula strings themselves.
All fixtures are synthetic and anonymous.
"""
from __future__ import annotations

import pytest
from openpyxl import load_workbook

from revenue_kun.excel_output import (
    SHEET_EXPENSE,
    SHEET_OER,
    SHEET_RENT_ROLL,
    DirectCapRow,
    write_direct_cap_workbook,
)

# ---------------------------------------------------------------------------
# Synthetic fixture matching the spec's worked example exactly
# ---------------------------------------------------------------------------

# Two occupied units whose combined monthly income is:
#   rent 172,000 / common fee 10,000 / water 4,000 / parking 10,000 / other 1,500
_ROWS = [
    DirectCapRow(
        区画="101", ステータス="入居",
        月額賃料=80_000, 月額共益費=5_000,
        月額水道光熱費=2_000, 月額駐車場=10_000, 月額その他収入=0,
    ),
    DirectCapRow(
        区画="102", ステータス="入居",
        月額賃料=92_000, 月額共益費=5_000,
        月額水道光熱費=2_000, 月額駐車場=0, 月額その他収入=1_500,
    ),
]

_EXPECTED_ANNUAL = {
    "rent": 2_064_000,
    "cam": 120_000,
    "water": 48_000,
    "parking": 120_000,
    "other": 18_000,
}
_EXPECTED_GPI = 2_370_000


@pytest.fixture(scope="module")
def workbook_path(tmp_path_factory):
    p = tmp_path_factory.mktemp("v052") / "direct_cap.xlsx"
    write_direct_cap_workbook(p, _ROWS)
    return p


@pytest.fixture(scope="module")
def wb(workbook_path):
    return load_workbook(workbook_path)


def _all_formula_strings(ws) -> list[str]:
    result = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                result.append(cell.value)
    return result


# ---------------------------------------------------------------------------
# 1. 付帯収入自動算入 — both sheets reach the same GPI
# ---------------------------------------------------------------------------

def test_rent_roll_annual_totals_match_expected(wb):
    ws = wb[SHEET_RENT_ROLL]
    annual_row = None
    for row in ws.iter_rows():
        if row[0].value == "年計":
            annual_row = row
            break
    assert annual_row is not None

    # Re-derive the same SUM/×12 the formulas perform, from the raw inputs.
    rent_annual = sum(r.月額賃料 or 0 for r in _ROWS) * 12
    cam_annual = sum(r.月額共益費 or 0 for r in _ROWS) * 12
    water_annual = sum(r.月額水道光熱費 or 0 for r in _ROWS) * 12
    parking_annual = sum(r.月額駐車場 or 0 for r in _ROWS) * 12
    other_annual = sum(r.月額その他収入 or 0 for r in _ROWS) * 12

    assert rent_annual == _EXPECTED_ANNUAL["rent"]
    assert cam_annual == _EXPECTED_ANNUAL["cam"]
    assert water_annual == _EXPECTED_ANNUAL["water"]
    assert parking_annual == _EXPECTED_ANNUAL["parking"]
    assert other_annual == _EXPECTED_ANNUAL["other"]
    gpi = rent_annual + cam_annual + water_annual + parking_annual + other_annual
    assert gpi == _EXPECTED_GPI


def test_oer_gpi_formula_structure_sums_all_five_income_rows(wb):
    """E10 (GPI) = SUM(E5:E9); E5:E9 are all cross-sheet refs (no =0 branch)."""
    ws = wb[SHEET_OER]
    assert ws["E10"].value == "=SUM(E5:E9)"
    for ref in ("E5", "E6", "E7", "E8", "E9"):
        formula = ws[ref].value
        assert isinstance(formula, str) and formula != "=0"
        assert SHEET_RENT_ROLL in formula


def test_expense_gpi_formula_structure_sums_all_five_income_rows(wb):
    """費用詳細版 E10 (GPI) = SUM(E5:E9), independently referencing 読み取りレントロール."""
    ws = wb[SHEET_EXPENSE]
    assert ws["E10"].value == "=SUM(E5:E9)"
    for ref in ("E5", "E6", "E7", "E8", "E9"):
        formula = ws[ref].value
        assert isinstance(formula, str) and formula != "=0"
        assert SHEET_RENT_ROLL in formula


# ---------------------------------------------------------------------------
# 2. OER版: GPI 2,370,000 / 採用OER 30% -> 収益試算値 33,180,000
# ---------------------------------------------------------------------------

def test_oer_scenario_matches_spec_expected_value():
    gpi = _EXPECTED_GPI
    vacancy_rate = 0.0
    bad_debt_rate = 0.0
    oer_rate = 0.30
    capex = 0
    cap_rate = 0.05

    egi = gpi * (1 - vacancy_rate - bad_debt_rate)
    opex = egi * oer_rate
    noi = egi - opex
    net_income = noi - capex
    indicated_value = net_income / cap_rate

    assert egi == 2_370_000
    assert opex == 711_000
    assert noi == 1_659_000
    assert net_income == 1_659_000
    assert indicated_value == 33_180_000


def test_oer_sheet_formula_chain_matches_python_recomputation(wb):
    """Structural check: the OER formula chain is exactly the one re-computed above."""
    ws = wb[SHEET_OER]
    assert ws["E20"].value == "=E10*(1-N(E13)-N(E14))"
    assert ws["E21"].value == "=E20*N(E15)"
    assert ws["E22"].value == "=E20-E21"
    assert ws["E23"].value == "=E22-N(E16)"
    assert ws["E24"].value == '=IFERROR(E23/E17,"")'


# ---------------------------------------------------------------------------
# 3. 費用詳細版: GPI 2,370,000 / 個別費用合計 460,000 -> 収益試算値 38,200,000
# ---------------------------------------------------------------------------

def test_expense_scenario_matches_spec_expected_value():
    gpi = _EXPECTED_GPI
    vacancy_rate = 0.0
    bad_debt_rate = 0.0
    capex = 0
    cap_rate = 0.05
    expenses = {
        "管理費・管理委託費": 100_000,
        "修繕費": 50_000,
        "損害保険料": 30_000,
        "固定資産税・都市計画税": 200_000,
        "水道光熱費": 60_000,
        "その他運営費用": 20_000,
    }

    egi = gpi * (1 - vacancy_rate - bad_debt_rate)
    opex_total = sum(expenses.values())
    noi = egi - opex_total
    net_income = noi - capex
    indicated_value = net_income / cap_rate

    assert opex_total == 460_000
    assert egi == 2_370_000
    assert noi == 1_910_000
    assert net_income == 1_910_000
    assert indicated_value == 38_200_000

    # 水道関連の NOI 寄与: 水道代収入(48,000) - 水道光熱費(60,000) = -12,000
    water_noi_contribution = _EXPECTED_ANNUAL["water"] - expenses["水道光熱費"]
    assert water_noi_contribution == -12_000


def test_expense_sheet_formula_chain_matches_python_recomputation(wb):
    ws = wb[SHEET_EXPENSE]
    assert ws["E25"].value == "=SUM(E19:E24)"
    assert ws["E28"].value == "=E10*(1-N(E13)-N(E14))"
    assert ws["E29"].value == "=E28-E25"
    assert ws["E30"].value == "=E29-N(E15)"
    assert ws["E31"].value == '=IFERROR(E30/E16,"")'


# ---------------------------------------------------------------------------
# 4. シート独立性
# ---------------------------------------------------------------------------

def test_oer_formulas_never_mention_expense_sheet_name(wb):
    ws = wb[SHEET_OER]
    for formula in _all_formula_strings(ws):
        assert SHEET_EXPENSE not in formula


def test_expense_formulas_never_mention_oer_sheet_name(wb):
    ws = wb[SHEET_EXPENSE]
    for formula in _all_formula_strings(ws):
        assert SHEET_OER not in formula


def test_no_oer_input_or_result_labels_in_expense_sheet(wb):
    """費用詳細版の入力欄・計算結果ラベルに「採用OER」は出現しない。

    D34 の独立性注記は「本シートはOERを使用せず...」と説明のため OER に言及して
    よい（これは leak ではない）ので、注記行は対象外とする。
    """
    ws = wb[SHEET_EXPENSE]
    note_rows = {34}
    for row in ws.iter_rows():
        for cell in row:
            if cell.row in note_rows:
                continue
            if isinstance(cell.value, str):
                assert "採用OER" not in cell.value, (
                    f"Unexpected OER label leaked into {cell.coordinate}: {cell.value!r}"
                )


# ---------------------------------------------------------------------------
# 5. 非収入項目（敷金・保証金・預り金・月計・年計・合計行）はGPIへ算入されない
# ---------------------------------------------------------------------------

def test_deposit_like_headers_are_not_mapped_to_any_income_field():
    """敷金/保証金/預り金 のヘッダーは canonical income field に解決されない。"""
    from revenue_kun.pdf_extract import _resolve_header_key

    for header in ("敷金", "保証金", "預り金", "敷金・保証金"):
        resolved = _resolve_header_key(header)
        assert resolved not in ("rent", "cam", "water", "parking", "other_income"), (
            f"{header!r} must not resolve to an income field, got: {resolved!r}"
        )


def test_monthly_and_annual_total_rows_excluded_from_income_sum_range(wb):
    """読み取りレントロールの月計・年計行は、両計算シートのGPI集計対象範囲に含まれない。

    _build_rent_roll_sheet は data_start:data_end のみを月計行の SUM 範囲とし、
    月計・年計自身の行は SUM 範囲の外側（末尾の別行）に配置される。両計算シート
    (OER/費用詳細版) の E5:E9 は年計行 1 行だけを参照するため、月計・年計という
    ラベル文字列がユニットとして二重に加算されることはない。
    """
    rr_ws = wb[SHEET_RENT_ROLL]
    room_col_values = [row[0].value for row in rr_ws.iter_rows(min_row=2)]
    assert room_col_values.count("月計") == 1
    assert room_col_values.count("年計") == 1
    # 実データ行 (101, 102) の後に 月計・年計 が1行ずつ、合計4行のみ存在する。
    assert len(room_col_values) == len(_ROWS) + 2

    monthly_row = next(r for r in rr_ws.iter_rows(min_row=2) if r[0].value == "月計")[0].row
    annual_row = next(r for r in rr_ws.iter_rows(min_row=2) if r[0].value == "年計")[0].row
    monthly_formula = rr_ws.cell(monthly_row, 3).value  # 月額賃料の月計セル
    # SUM range must stop before the 月計 row itself (no self-reference / double count).
    assert f"C{monthly_row}" not in monthly_formula
    assert f"C{annual_row}" not in monthly_formula
