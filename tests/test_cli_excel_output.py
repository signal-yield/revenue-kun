"""CLI integration tests for --excel-output flag.

All fixtures are synthetic and anonymous.  No private PDFs or PII are used.
"""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from revenue_kun.cli import run
from revenue_kun.excel_output import SHEET_EXPENSE, SHEET_OER, SHEET_RENT_ROLL

ASSUMPTIONS = "assumptions.sample.yaml"
DUMMY_CSV = "data/dummy_rent_roll.csv"


# ---------------------------------------------------------------------------
# --excel-output creates the workbook
# ---------------------------------------------------------------------------

def test_excel_output_flag_creates_file(tmp_path):
    """--excel-output パスを渡すと .xlsx が生成される。"""
    xlsx = tmp_path / "out" / "direct_cap.xlsx"
    rc = run(ASSUMPTIONS, DUMMY_CSV, str(tmp_path / "out"),
             excel_output_path=str(xlsx))
    assert rc == 0
    assert xlsx.exists(), f"Expected {xlsx} to be created"


def test_excel_output_has_all_three_sheets(tmp_path):
    """生成されたワークブックに3シートすべてが含まれる。"""
    xlsx = tmp_path / "out" / "direct_cap.xlsx"
    run(ASSUMPTIONS, DUMMY_CSV, str(tmp_path / "out"),
        excel_output_path=str(xlsx))
    wb = load_workbook(xlsx)
    assert SHEET_OER in wb.sheetnames, f"{SHEET_OER!r} missing"
    assert SHEET_EXPENSE in wb.sheetnames, f"{SHEET_EXPENSE!r} missing"
    assert SHEET_RENT_ROLL in wb.sheetnames, f"{SHEET_RENT_ROLL!r} missing"


def test_excel_output_sheet_order(tmp_path):
    """シート順: OER / 費用詳細版 / レントロール。"""
    xlsx = tmp_path / "out" / "direct_cap.xlsx"
    run(ASSUMPTIONS, DUMMY_CSV, str(tmp_path / "out"),
        excel_output_path=str(xlsx))
    wb = load_workbook(xlsx)
    names = wb.sheetnames
    assert names[0] == SHEET_OER
    assert names[1] == SHEET_EXPENSE
    assert names[2] == SHEET_RENT_ROLL


def test_excel_output_path_shown_in_stdout(tmp_path, capsys):
    """--excel-output パスが「出力ファイル:」ブロックに出力される。"""
    xlsx = tmp_path / "out" / "direct_cap.xlsx"
    run(ASSUMPTIONS, DUMMY_CSV, str(tmp_path / "out"),
        excel_output_path=str(xlsx))
    out = capsys.readouterr().out
    assert str(xlsx) in out


# ---------------------------------------------------------------------------
# Without --excel-output, existing behavior is unchanged
# ---------------------------------------------------------------------------

def test_no_excel_output_flag_no_extra_file(tmp_path):
    """--excel-output 未指定では余分な .xlsx を生成しない。"""
    out_dir = tmp_path / "out"
    run(ASSUMPTIONS, DUMMY_CSV, str(out_dir))
    # The standard revenue_analysis.xlsx is expected; direct_cap.xlsx is not.
    assert not (out_dir / "direct_cap.xlsx").exists()


def test_existing_outputs_still_created_with_excel_flag(tmp_path):
    """--excel-output 指定時も既存の出力ファイルは生成される。"""
    out_dir = tmp_path / "out"
    xlsx = out_dir / "direct_cap.xlsx"
    rc = run(ASSUMPTIONS, DUMMY_CSV, str(out_dir),
             excel_output_path=str(xlsx))
    assert rc == 0
    assert (out_dir / "revenue_analysis.xlsx").exists()
    assert (out_dir / "missing_info.md").exists()
    assert (out_dir / "extraction_log.json").exists()


# ---------------------------------------------------------------------------
# --dry-run suppresses --excel-output
# ---------------------------------------------------------------------------

def test_dry_run_suppresses_excel_output(tmp_path):
    """--dry-run では --excel-output を指定しても .xlsx が生成されない。"""
    out_dir = tmp_path / "out"
    xlsx = out_dir / "direct_cap.xlsx"
    rc = run(ASSUMPTIONS, DUMMY_CSV, str(out_dir),
             dry_run=True, excel_output_path=str(xlsx))
    assert rc == 0
    assert not xlsx.exists(), "dry-run should not write the Excel workbook"


# ---------------------------------------------------------------------------
# Workbook content: rent roll rows reflect DUMMY_CSV units
# ---------------------------------------------------------------------------

def test_rent_roll_sheet_has_data_rows(tmp_path):
    """読み取りレントロールシートに少なくとも1行のデータがある。"""
    xlsx = tmp_path / "out" / "direct_cap.xlsx"
    run(ASSUMPTIONS, DUMMY_CSV, str(tmp_path / "out"),
        excel_output_path=str(xlsx))
    wb = load_workbook(xlsx)
    ws = wb[SHEET_RENT_ROLL]
    # Row 1 is the header; at least row 2 should have unit data.
    assert ws.max_row >= 2, "Rent roll sheet should have at least one data row"
    # First column of row 2 should be non-empty (区画番号)
    assert ws.cell(2, 1).value is not None, "Row 2 col A should contain a unit number"
