"""skill パッケージの self-test。

カバー範囲:
  1. skill/scripts/ が src/ とバイト一致（エンジン同期チェック）
  2. Skill エンジン単体で dry-run が exit 0
  3. Skill エンジン単体で xlsx を生成できる
  4. OER 自己計算モデル検証（E24=IFERROR、判断値セル E13:E17 が空欄）
  5. SKILL.md 免責ブロック存在確認
  6. SKILL.md トリガー衝突語チェック（土地査定/重説/DCF が非対象明記以外にない）
"""
from __future__ import annotations

import filecmp
import subprocess
import sys
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parents[2]
SKILL_DIR = ROOT / "skill"
SKILL_SCRIPTS = SKILL_DIR / "scripts"
SKILL_PKG = SKILL_SCRIPTS / "revenue_kun"
SRC_PKG = ROOT / "src" / "revenue_kun"
SAMPLES = SKILL_DIR / "samples"
SIMPLE_PDF = SAMPLES / "sample_rentroll_simple.pdf"
ASSUMPTIONS = SAMPLES / "assumptions.sample.yaml"
OUT_DIR = SKILL_DIR / "out"


# ---------------------------------------------------------------------------
# 1. エンジン同期チェック
# ---------------------------------------------------------------------------

def test_skill_scripts_in_sync_with_src():
    """skill/scripts/revenue_kun/ が src/revenue_kun/ とバイト一致する。"""
    assert SKILL_PKG.exists(), "skill/scripts/revenue_kun/ が存在しません。python build_skill.py を実行してください。"
    result = filecmp.dircmp(str(SRC_PKG), str(SKILL_PKG))

    def collect_diffs(cmp: filecmp.dircmp) -> list[str]:
        diffs = list(cmp.diff_files) + list(cmp.left_only) + list(cmp.right_only)
        for sub in cmp.subdirs.values():
            diffs.extend(collect_diffs(sub))
        return diffs

    diffs = collect_diffs(result)
    assert not diffs, (
        f"skill/scripts/revenue_kun/ と src/revenue_kun/ が一致しません: {diffs}\n"
        "python build_skill.py を実行して再生成してください。"
    )


# ---------------------------------------------------------------------------
# 2. dry-run
# ---------------------------------------------------------------------------

def test_skill_dryrun_exit_zero(tmp_path):
    """Skill エンジン単体の --dry-run が exit 0 で完了する。"""
    result = subprocess.run(
        [
            sys.executable,
            str(SKILL_SCRIPTS / "main.py"),
            "--assumptions", str(ASSUMPTIONS),
            "--rent-roll-pdf", str(SIMPLE_PDF),
            "--output", str(tmp_path),
            "--dry-run",
        ],
        capture_output=True,
        text=True,
    )
    assert result.returncode == 0, f"dry-run failed:\n{result.stderr}"


# ---------------------------------------------------------------------------
# 3 & 4. フルラン + OER モデル検証
# ---------------------------------------------------------------------------

def test_skill_full_run_and_oer_model(tmp_path):
    """Skill エンジンが xlsx を生成し、OER 自己計算モデルが正しい。"""
    xlsx_path = tmp_path / "test_output.xlsx"
    result = subprocess.run(
        [
            sys.executable,
            str(SKILL_SCRIPTS / "main.py"),
            "--assumptions", str(ASSUMPTIONS),
            "--rent-roll-pdf", str(SIMPLE_PDF),
            "--output", str(tmp_path),
            "--excel-output", str(xlsx_path),
        ],
        capture_output=True,
        text=True,
    )
    assert result.returncode == 0, f"full run failed:\n{result.stderr}"
    assert xlsx_path.exists(), "xlsx が生成されていません"

    from openpyxl import load_workbook
    wb = load_workbook(str(xlsx_path))
    ws = wb["直接還元法_OER"]

    # E24 = 収益試算値セル（IFERROR 式）
    assert ws["E24"].value == '=IFERROR(E23/E17,"")', f"E24 が期待する式ではありません: {ws['E24'].value}"

    # 判断値セル（E13:E17）は空欄
    for cell_ref in ("E13", "E14", "E15", "E16", "E17"):
        assert ws[cell_ref].value is None, f"{cell_ref} は空欄であるべきですが値があります: {ws[cell_ref].value}"


# ---------------------------------------------------------------------------
# 5. SKILL.md 免責ブロック
# ---------------------------------------------------------------------------

REQUIRED_DISCLAIMER_PHRASES = [
    "収益試算値",
    "収益価格」ではありません",
    "Issue #21 open",
    "実務検証済みとは表記しません",
    "補完しません",
]


def test_skill_md_has_disclaimer():
    """SKILL.md に免責ブロックの必須フレーズが含まれる。"""
    skill_md = (SKILL_DIR / "SKILL.md").read_text(encoding="utf-8")
    for phrase in REQUIRED_DISCLAIMER_PHRASES:
        assert phrase in skill_md, f"SKILL.md に必須フレーズがありません: {phrase!r}"


# ---------------------------------------------------------------------------
# 6. トリガー衝突語チェック
# ---------------------------------------------------------------------------

COLLISION_PATTERNS = ["土地査定", "査定書", "重説", "重調", "35条", "DCF"]
NEGATION_MARKERS = ["非対象", "対象外", "→"]


def test_skill_md_trigger_no_collision():
    """SKILL.md の衝突語（土地査定/重説/DCF 等）が非対象明記以外に現れない。"""
    skill_md = (SKILL_DIR / "SKILL.md").read_text(encoding="utf-8")
    violations: list[str] = []
    for line in skill_md.splitlines():
        for pat in COLLISION_PATTERNS:
            if pat in line and not any(neg in line for neg in NEGATION_MARKERS):
                violations.append(f"{pat!r} in: {line.strip()}")
    assert not violations, f"SKILL.md にトリガー衝突語が含まれています:\n" + "\n".join(violations)
