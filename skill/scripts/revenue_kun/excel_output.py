"""Excel workbook generation for direct capitalization output.

Produces a three-sheet .xlsx workbook:

  Sheet 1: 直接還元法_OER        -- self-computing live model; income E5:E9
                                    linked to 読み取りレントロール; EGI/NOI/
                                    収益試算値 computed from user inputs E13:E17
  Sheet 2: 直接還元法‗費用詳細版 -- independent self-computing live model;
                                    income E5:E9 linked to 読み取りレントロール
                                    (its own reference, not the OER sheet's);
                                    EGI/NOI/収益試算値 computed from user
                                    inputs (vacancy/bad-debt/capex/cap rate)
                                    and individual expense line items
  Sheet 3: 読み取りレントロール    -- extracted rent roll with monthly+annual totals

v0.5.2 product boundary (confirmed):
  revenue-kun's job stops at generating the 3-sheet workbook. The app does
  not ask the user to choose between the OER sheet and the expense-detail
  sheet, and does not collect 用途区分 / OER / 空室損失率 / 貸倒損失率 /
  個別費用 / 資本的支出 / 還元利回り anywhere in the CLI or Web UI. Both
  sheets are always generated, are fully independent of each other (neither
  references the other's input cells, expense totals, NOI, or 収益試算値),
  and the user decides after receiving the Excel file which sheet(s) to use
  and what to type into the empty input cells.

収入の自動算入（v0.5.2 で確定）:
  賃料・共益費・水道代収入・駐車場収入・その他収入は、いずれも常に両計算
  シートの収入項目としてクロスシート参照される。「算入対象外」というopt-out
  表現やopt-in選択は廃止した。読み取りレントロールに表示される値がそのまま
  両シートの GPI 集計に反映される（欠損値は 0 として合算される — 読み取り
  レントロール側の SUM 式が空欄を 0 として扱うため、生成ロジック側での特別
  な補完は不要）。

IMPORTANT: income cells reference the annual total row in 読み取りレントロール
directly.  They do NOT multiply by 12 because 読み取りレントロール already
provides annual totals (monthly_total * 12).

OER calculation chain (all Excel-side):
  E10 = SUM(E5:E9)                   GPI (gross potential income)
  E20 = E10*(1-N(E13)-N(E14))        EGI (effective gross income)
  E21 = E20*N(E15)                   operating expenses (EGI × 採用OER)
  E22 = E20-E21                      NOI
  E23 = E22-N(E16)                   net income (after capex)
  E24 = IFERROR(E23/E17,"")          収益試算値 — empty when E17 is blank

直接還元法‗費用詳細版 calculation chain (independent of OER, all Excel-side):
  E10 = SUM(E5:E9)                   GPI
  E25 = SUM(E19:E24)                 operating expenses (individual line items)
  E28 = E10*(1-N(E13)-N(E14))        EGI
  E29 = E28-E25                      NOI
  E30 = E29-N(E15)                   net income (after capex)
  E31 = IFERROR(E30/E16,"")          収益試算値 — empty when E16 is blank

Sheet independence (v0.5.2): neither sheet's formulas reference the other
sheet by name. Both reference only 読み取りレントロール for their income
block; everything downstream of that is computed independently within each
sheet.
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .config import OptionalIncomeConfig
from .rent_roll import RentRollUnit

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_VACANT_NOTE = "ユーザーが賃料等を入力可能"
_MONEY_FORMAT = "#,##0"

# Exact sheet names.  費用詳細版 uses U+2017 DOUBLE LOW LINE, not U+005F underscore.
SHEET_OER = "直接還元法_OER"
SHEET_EXPENSE = "直接還元法‗費用詳細版"
SHEET_RENT_ROLL = "読み取りレントロール"

# Column positions in 読み取りレントロール (1-based, openpyxl convention)
_C_ROOM = 1     # 部屋番号
_C_STATUS = 2   # ステータス
_C_RENT = 3     # 月額賃料
_C_CAM = 4      # 月額共益費
_C_UTIL = 5     # 月額水道光熱費
_C_PARKING = 6  # 月額駐車場
_C_OTHER = 7    # 月額その他収入
_C_NOTES = 8    # 備考
_INCOME_COLS = (_C_RENT, _C_CAM, _C_UTIL, _C_PARKING, _C_OTHER)

# Income rows shared by both calculation sheets: (row_num, label, income_col_in_rent_roll).
# All five are always cross-sheet references — there is no opt-in/opt-out
# branch (removed in v0.5.2; see module docstring).
_INCOME_ROWS: list[tuple[int, str, int]] = [
    (5, "貸室賃料収入", _C_RENT),
    (6, "共益費収入",   _C_CAM),
    (7, "水道代収入",   _C_UTIL),
    (8, "駐車場収入",   _C_PARKING),
    (9, "その他収入",   _C_OTHER),
]

# Styles
_THIN = Side(style="thin")
_FULL_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL = PatternFill("solid", fgColor="305496")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_BOLD = Font(bold=True)
_WARN_FONT = Font(color="C00000")
_INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")   # light yellow for user-input cells


# ---------------------------------------------------------------------------
# Public dataclass
# ---------------------------------------------------------------------------

@dataclass
class DirectCapRow:
    """One unit row for direct-capitalization Excel output.

    Covers the five income columns required by the spec (賃料/共益費/
    水道光熱費/駐車場/その他収入).  The latter three default to None
    because the current RentRollUnit does not track them; users fill
    them in Excel.
    """

    区画: str
    ステータス: str | None
    月額賃料: float | None
    月額共益費: float | None
    月額水道光熱費: float | None = None
    月額駐車場: float | None = None
    月額その他収入: float | None = None
    備考: str | None = None

    @classmethod
    def from_rent_roll_unit(cls, unit: RentRollUnit) -> DirectCapRow:
        """Convert a RentRollUnit to a DirectCapRow.

        All income fields (賃料/共益費/水道代収入/駐車場収入/その他収入) are
        always populated from the unit's extracted values. They flow into
        読み取りレントロール, and from there — unconditionally — into both
        calculation sheets' income blocks (see write_direct_cap_workbook).
        Vacant units receive the standard 備考 note automatically.
        """
        note = _VACANT_NOTE if not unit.is_occupied else None
        return cls(
            区画=unit.区画,
            ステータス=unit.稼働状況,
            月額賃料=unit.月額賃料_円,
            月額共益費=unit.月額共益費_円,
            月額水道光熱費=unit.月額水道代_円,
            月額駐車場=unit.月額駐車場収入_円,
            月額その他収入=unit.月額その他収入_円,
            備考=note,
        )


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def write_direct_cap_workbook(
    path: str | Path,
    rows: list[DirectCapRow],
    oi_config: OptionalIncomeConfig | None = None,
) -> None:
    """Write the three-sheet direct-capitalization workbook to *path*.

    Sheet order: 直接還元法_OER / 直接還元法‗費用詳細版 / 読み取りレントロール

    読み取りレントロール: 抽出した付帯収入（水道代収入/駐車場収入/その他収入）を
    常に表示する。

    直接還元法_OER と 直接還元法‗費用詳細版: いずれも読み取りレントロールの
    年計行から収入（賃料・共益費・水道代収入・駐車場収入・その他収入）を
    無条件にクロスシート参照し、GPI へ算入する。両シートは互いに独立して
    おり、一方の入力値・計算値・シートをもう一方が参照することはない。

    *oi_config* は後方互換のためのみ受理するパラメータです（deprecated）。
    v0.5.1 以前の ``assumptions.yaml`` の ``optional_income.include_in_gpi`` /
    ``columns`` 設定、および Web UI の旧チェックボックス選択はこの引数として
    渡され得ますが、v0.5.2 以降は計算結果に一切影響しません。収入は常に
    自動算入されます。呼び出し側（CLI / Web UI）はこの引数を省略しても
    同じ結果になります。

    OER cells E13:E17 are empty user-input cells (border + light-yellow fill).
    OER cells E20:E24 compute EGI → expenses → NOI → 収益試算値 automatically.

    費用詳細版 cells E13:E16（前提条件）と E19:E24（運営費用明細）は空欄の
    ユーザー入力セル。E28:E31 が EGI → NOI → 純収益 → 収益試算値 を自動計算する。
    """
    # oi_config is accepted for backward compatibility only; it is
    # intentionally unused. See docstring above and README for the v0.5.2
    # deprecation note.
    del oi_config

    wb = Workbook()

    # Build the rent roll sheet first to know the annual_total_row index.
    rr_ws = wb.active
    rr_ws.title = SHEET_RENT_ROLL
    annual_total_row = _build_rent_roll_sheet(rr_ws, rows)

    # Insert OER as sheet 0 (leftmost tab).
    oer_ws = wb.create_sheet(SHEET_OER, 0)
    _build_oer_sheet(oer_ws, annual_total_row)

    # Insert expense detail as sheet 1 (between OER and rent roll).
    exp_ws = wb.create_sheet(SHEET_EXPENSE, 1)
    _build_expense_sheet(exp_ws, annual_total_row)

    # Remove the default empty sheet created by Workbook().
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    dest = Path(path)
    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)


# ---------------------------------------------------------------------------
# Sheet builders (private)
# ---------------------------------------------------------------------------

def _build_rent_roll_sheet(ws, rows: list[DirectCapRow]) -> int:
    """Populate 読み取りレントロール.  Returns the annual_total_row (1-based)."""
    headers = [
        "部屋番号", "ステータス",
        "月額賃料", "月額共益費", "水道代収入（月額）", "駐車場収入（月額）", "その他収入（月額）",
        "備考",
    ]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    data_start = 2
    for offset, row in enumerate(rows):
        r = data_start + offset
        note = row.備考
        if note is None and _is_vacant(row.ステータス):
            note = _VACANT_NOTE
        ws.cell(r, _C_ROOM,    row.区画)
        ws.cell(r, _C_STATUS,  row.ステータス or "")
        ws.cell(r, _C_RENT,    row.月額賃料)
        ws.cell(r, _C_CAM,     row.月額共益費)
        ws.cell(r, _C_UTIL,    row.月額水道光熱費)
        ws.cell(r, _C_PARKING, row.月額駐車場)
        ws.cell(r, _C_OTHER,   row.月額その他収入)
        ws.cell(r, _C_NOTES,   note)
        for c in _INCOME_COLS:
            ws.cell(r, c).number_format = _MONEY_FORMAT

    n = len(rows)
    data_end = data_start + n - 1  # last data row (inclusive)
    monthly_row = data_start + n       # row after last data row
    annual_row = monthly_row + 1

    # Monthly total row
    ws.cell(monthly_row, _C_ROOM, "月計")
    ws.cell(monthly_row, _C_NOTES, "月額合計")
    for c in _INCOME_COLS:
        col_letter = get_column_letter(c)
        if n > 0:
            formula = f"=SUM({col_letter}{data_start}:{col_letter}{data_end})"
        else:
            formula = "=0"
        cell = ws.cell(monthly_row, c, formula)
        cell.number_format = _MONEY_FORMAT

    # Annual total row  (monthly * 12, with full border)
    ws.cell(annual_row, _C_ROOM, "年計")
    ws.cell(annual_row, _C_NOTES, "年額合計")
    for c in _INCOME_COLS:
        col_letter = get_column_letter(c)
        monthly_cell_ref = f"{col_letter}{monthly_row}"
        formula = f"={monthly_cell_ref}*12"
        cell = ws.cell(annual_row, c, formula)
        cell.number_format = _MONEY_FORMAT
        cell.border = _FULL_BORDER

    return annual_row


def _write_income_block(ws, annual_total_row: int) -> None:
    """Write the shared income block (D4:E10) cross-referencing 読み取りレントロール.

    Used identically by both 直接還元法_OER and 直接還元法‗費用詳細版 so that
    each sheet independently derives the same GPI from 読み取りレントロール,
    without referencing each other.
    """
    ann = annual_total_row

    ws["D4"] = "収入項目"
    ws["E4"] = "年額（円）"
    ws["D4"].font = _BOLD
    ws["E4"].font = _BOLD

    for row_num, label, income_col in _INCOME_ROWS:
        col_letter = get_column_letter(income_col)
        ws.cell(row_num, 4, label)
        formula = f"='{SHEET_RENT_ROLL}'!{col_letter}{ann}"
        cell = ws.cell(row_num, 5, formula)
        cell.number_format = _MONEY_FORMAT

    ws["D10"] = "総収入（年額）"
    ws["D10"].font = _BOLD
    ws["E10"] = "=SUM(E5:E9)"
    ws["E10"].number_format = _MONEY_FORMAT
    ws["E10"].font = _BOLD


def _write_input_cell(ws, row: int, label: str, fmt: str) -> None:
    """Write one empty, visually-marked user-input cell at D{row}/E{row}."""
    ws.cell(row, 4, label)
    cell = ws.cell(row, 5)
    cell.number_format = fmt
    cell.border = _FULL_BORDER
    cell.fill = _INPUT_FILL


def _build_oer_sheet(ws, annual_total_row: int) -> None:
    """Populate 直接還元法_OER as a self-computing live model.

    Income block (E5:E9): always cross-sheet refs to 読み取りレントロール
    annual row (see _write_income_block). No opt-in/opt-out branch.

    Input cells (E13:E17): empty with border + light-yellow fill.
    User enters: 空室損失率 / 貸倒損失率 / 採用OER / 資本的支出 / 還元利回り.

    Calculation chain (E20:E24):
      E20 EGI  = E10*(1-N(E13)-N(E14))
      E21 opex = E20*N(E15)          採用OER（運営費用 ÷ EGI）× EGI
      E22 NOI  = E20-E21
      E23 net  = E22-N(E16)
      E24 value= IFERROR(E23/E17,"") (empty when cap rate is blank)

    This sheet does not reference 直接還元法‗費用詳細版 in any way (v0.5.2
    sheet-independence requirement).
    """
    # Title and disclaimer
    ws["A1"] = "直接還元法（収益試算） OER方式"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "※ 本値は「収益試算値」であり「収益価格」ではありません。"
    ws["A2"].font = _WARN_FONT

    _write_income_block(ws, annual_total_row)

    # Assumption inputs header
    ws["D12"] = "前提条件（ユーザー入力）"
    ws["D12"].font = _BOLD

    _write_input_cell(ws, 13, "空室損失率", "0.0%")
    _write_input_cell(ws, 14, "貸倒損失率", "0.0%")
    _write_input_cell(ws, 15, "採用OER（運営費用 ÷ EGI）", "0.0%")
    _write_input_cell(ws, 16, "資本的支出（年額）", _MONEY_FORMAT)
    _write_input_cell(ws, 17, "還元利回り", "0.000%")

    # Calculation results header
    ws["D19"] = "計算結果"
    ws["D19"].font = _BOLD

    ws["D20"] = "有効総収入（EGI）"
    ws["E20"] = "=E10*(1-N(E13)-N(E14))"
    ws["E20"].number_format = _MONEY_FORMAT

    ws["D21"] = "運営費用（EGI×採用OER）"
    ws["E21"] = "=E20*N(E15)"
    ws["E21"].number_format = _MONEY_FORMAT

    ws["D22"] = "運営純収益（NOI）"
    ws["E22"] = "=E20-E21"
    ws["E22"].number_format = _MONEY_FORMAT

    ws["D23"] = "純収益（還元対象）"
    ws["E23"] = "=E22-N(E16)"
    ws["E23"].number_format = _MONEY_FORMAT

    ws["D24"] = "収益試算値（直接還元法）"
    ws["D24"].font = _BOLD
    ws["E24"] = '=IFERROR(E23/E17,"")'
    ws["E24"].number_format = _MONEY_FORMAT
    ws["E24"].font = _BOLD

    ws["D25"] = "※ 還元利回り（E17）未入力時、収益試算値は空欄になります。"
    ws["D25"].font = _WARN_FONT

    ws["D27"] = (
        "※ 本シートは個別の運営費用明細が不明な場合の簡便法です（運営費用＝EGI×採用OER）。"
        "個別費用が判明している場合は 直接還元法‗費用詳細版 をご利用ください。"
        "本シートは 直接還元法‗費用詳細版 の入力値・計算値を参照せず、独立して計算します。"
    )
    ws["D27"].font = _WARN_FONT


def _build_expense_sheet(ws, annual_total_row: int) -> None:
    """Populate 直接還元法‗費用詳細版 as an independent self-computing live model.

    Income block (E5:E10): identical structure to the OER sheet, but its own
    formulas referencing 読み取りレントロール directly — this sheet never
    reads 直接還元法_OER.

    Input cells: 空室損失率/貸倒損失率/資本的支出/還元利回り (E13:E16) and the
    individual expense line items (E19:E24), all empty with border +
    light-yellow fill. The app does not estimate or pre-fill any of these.

    Calculation chain:
      E25 opex = SUM(E19:E24)         individual expense line items
      E28 EGI  = E10*(1-N(E13)-N(E14))
      E29 NOI  = E28-E25
      E30 net  = E29-N(E15)
      E31 value= IFERROR(E30/E16,"") (empty when cap rate is blank)

    This sheet does not use OER (経費率) at all, and does not reference
    直接還元法_OER in any way (v0.5.2 sheet-independence requirement).
    """
    ws["A1"] = "直接還元法（収益試算） 費用詳細方式"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "※ 本値は「収益試算値」であり「収益価格」ではありません。"
    ws["A2"].font = _WARN_FONT

    _write_income_block(ws, annual_total_row)

    # Assumption inputs header (vacancy/bad-debt/capex/cap rate — no OER here)
    ws["D12"] = "前提条件（ユーザー入力）"
    ws["D12"].font = _BOLD

    _write_input_cell(ws, 13, "空室損失率", "0.0%")
    _write_input_cell(ws, 14, "貸倒損失率", "0.0%")
    _write_input_cell(ws, 15, "資本的支出（年額）", _MONEY_FORMAT)
    _write_input_cell(ws, 16, "還元利回り", "0.000%")

    # Individual expense line items
    ws["D18"] = "運営費用明細（ユーザー入力）"
    ws["D18"].font = _BOLD

    expense_labels = [
        (19, "管理費・管理委託費"),
        (20, "修繕費"),
        (21, "損害保険料"),
        (22, "固定資産税・都市計画税"),
        (23, "水道光熱費"),
        (24, "その他運営費用"),
    ]
    for row_num, label in expense_labels:
        _write_input_cell(ws, row_num, label, _MONEY_FORMAT)

    ws["D25"] = "運営費用合計"
    ws["D25"].font = _BOLD
    ws["E25"] = "=SUM(E19:E24)"
    ws["E25"].number_format = _MONEY_FORMAT
    ws["E25"].font = _BOLD

    # Calculation results
    ws["D27"] = "計算結果"
    ws["D27"].font = _BOLD

    ws["D28"] = "有効総収入（EGI）"
    ws["E28"] = "=E10*(1-N(E13)-N(E14))"
    ws["E28"].number_format = _MONEY_FORMAT

    ws["D29"] = "運営純収益（NOI）"
    ws["E29"] = "=E28-E25"
    ws["E29"].number_format = _MONEY_FORMAT

    ws["D30"] = "純収益（還元対象）"
    ws["E30"] = "=E29-N(E15)"
    ws["E30"].number_format = _MONEY_FORMAT

    ws["D31"] = "収益試算値（直接還元法）"
    ws["D31"].font = _BOLD
    ws["E31"] = '=IFERROR(E30/E16,"")'
    ws["E31"].number_format = _MONEY_FORMAT
    ws["E31"].font = _BOLD

    ws["D32"] = "※ 還元利回り（E16）未入力時、収益試算値は空欄になります。"
    ws["D32"].font = _WARN_FONT

    ws["D34"] = (
        "※ 本シートは個別の運営費用明細が判明している場合にご利用ください。"
        "費用の全項目が判明していることは利用条件ではありません（判明した項目のみ入力してください）。"
        "本シートはOERを使用せず、直接還元法_OER の入力値・計算値を参照せず、独立して計算します。"
    )
    ws["D34"].font = _WARN_FONT


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _is_vacant(status: str | None) -> bool:
    """Return True if *status* indicates a vacant unit."""
    s = (status or "").strip()
    return any(k in s for k in ("空室", "空き", "募集"))
