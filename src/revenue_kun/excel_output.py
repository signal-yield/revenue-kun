"""Excel workbook generation for direct capitalization output.

Produces a three-sheet .xlsx workbook:

  Sheet 1: 直接還元法_OER        -- self-computing live model; income E5:E9
                                    linked to 読み取りレントロール; EGI/NOI/
                                    収益試算値 computed from user inputs E13:E17
  Sheet 2: 直接還元法‗費用詳細版 -- expense detail input (user-editable; SUM at B10)
  Sheet 3: 読み取りレントロール    -- extracted rent roll with monthly+annual totals

「表示」と「GPI算入」の分離:
  読み取りレントロール: 抽出された付帯収入（水道代収入/駐車場収入/その他収入）を
    常に表示する。opt-in/opt-out に関わらず抽出値を記録する。
  直接還元法_OER: optional_income.include_in_gpi=True かつ columns に含まれる
    収入のみ E7-E9 を cross-sheet ref にする。opt-out 時は =0（算入対象外）。

IMPORTANT: OER cells E5:E9 reference the annual total row in
読み取りレントロール directly.  They do NOT multiply by 12 because
読み取りレントロール already provides annual totals (monthly_total * 12).

OER calculation chain (all Excel-side):
  E10 = SUM(E5:E9)                   GPI (gross potential income)
  E20 = E10*(1-N(E13)-N(E14))        EGI (effective gross income)
  E21 = E20*N(E15)                   operating expenses (EGI × expense ratio)
  E22 = E20-E21                      NOI
  E23 = E22-N(E16)                   net income (after capex)
  E24 = IFERROR(E23/E17,"")          収益試算値 — empty when E17 is blank
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .config import OptionalIncomeConfig
from .rent_roll import RentRollUnit

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_VACANT_NOTE = "ユーザーが賃料等を入力可能"
_MONEY_FORMAT = "#,##0"

# Exact sheet names.  費用詳細版 uses U+2017 DOUBLE LOW LINE, not U+005F underscore.
SHEET_OER = "直接還元法_OER"
SHEET_EXPENSE = "直接還元法‗費用詳細版"
SHEET_RENT_ROLL = "読み取りレントロール"

# Column positions in 読み取りレントロール (1-based, openpyxl convention)
_C_ROOM = 1     # 部屋番号
_C_STATUS = 2   # ステータス
_C_RENT = 3     # 月額賃料
_C_CAM = 4      # 月額共益費
_C_UTIL = 5     # 月額水道光熱費
_C_PARKING = 6  # 月額駐車場
_C_OTHER = 7    # 月額その他収入
_C_NOTES = 8    # 備考
_INCOME_COLS = (_C_RENT, _C_CAM, _C_UTIL, _C_PARKING, _C_OTHER)

# OER income rows: (oer_row, label, income_col_in_rent_roll, optional_key | None)
# optional_key が None の行（賃料/共益費）は常に cross-sheet ref。
# optional_key が文字列の行（付帯収入）は opt-in 時のみ cross-sheet ref、
# opt-out 時は =0 & ラベルに「（算入対象外）」を付ける。
# 「水道代収入」は費用サイドの「水道光熱費」と表記上も混同しない。
_OER_INCOME_ROWS: list[tuple[int, str, int, str | None]] = [
    (5, "貸室賃料収入", _C_RENT,    None),
    (6, "共益費収入",   _C_CAM,     None),
    (7, "水道代収入",   _C_UTIL,    "water"),
    (8, "駐車場収入",   _C_PARKING, "parking"),
    (9, "その他収入",   _C_OTHER,   "other_income"),
]

# Styles
_THIN = Side(style="thin")
_FULL_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL = PatternFill("solid", fgColor="305496")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_BOLD = Font(bold=True)
_WARN_FONT = Font(color="C00000")
_INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")   # light yellow for user-input cells


# ---------------------------------------------------------------------------
# Public dataclass
# ---------------------------------------------------------------------------

@dataclass
class DirectCapRow:
    """One unit row for direct-capitalization Excel output.

    Covers the five income columns required by the spec (賃料/共益費/
    水道光熱費/駐車場/その他収入).  The latter three default to None
    because the current RentRollUnit does not track them; users fill
    them in Excel.
    """

    区画: str
    ステータス: str | None
    月額賃料: float | None
    月額共益費: float | None
    月額水道光熱費: float | None = None
    月額駐車場: float | None = None
    月額その他収入: float | None = None
    備考: str | None = None

    @classmethod
    def from_rent_roll_unit(cls, unit: RentRollUnit) -> DirectCapRow:
        """Convert a RentRollUnit to a DirectCapRow.

        Optional income fields (水道代収入/駐車場収入/その他収入) are always
        populated from the unit's extracted values so they appear in
        読み取りレントロール regardless of opt-in/opt-out setting.
        GPI inclusion is controlled by OER sheet formulas (=0 when opt-out,
        cross-sheet ref when opt-in) — see write_direct_cap_workbook.
        Vacant units receive the standard 備考 note automatically.
        """
        note = _VACANT_NOTE if not unit.is_occupied else None
        return cls(
            区画=unit.区画,
            ステータス=unit.稼働状況,
            月額賃料=unit.月額賃料_円,
            月額共益費=unit.月額共益費_円,
            月額水道光熱費=unit.月額水道代_円,
            月額駐車場=unit.月額駐車場収入_円,
            月額その他収入=unit.月額その他収入_円,
            備考=note,
        )


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def write_direct_cap_workbook(
    path: str | Path,
    rows: list[DirectCapRow],
    oi_config: OptionalIncomeConfig | None = None,
) -> None:
    """Write the three-sheet direct-capitalization workbook to *path*.

    Sheet order: 直接還元法_OER / 直接還元法‗費用詳細版 / 読み取りレントロール

    読み取りレントロール: 抽出した付帯収入（水道代収入/駐車場収入/その他収入）を
    opt-in/opt-out に関わらず常に表示する。

    直接還元法_OER: oi_config で指定された付帯収入のみ cross-sheet ref で GPI に算入。
    opt-out の付帯収入行は =0 を書き込み「（算入対象外）」と表示する。

    OER cells E13:E17 are empty user-input cells (border + light-yellow fill).
    OER cells E20:E24 compute EGI → expenses → NOI → 収益試算値 automatically.
    """
    _oi = oi_config or OptionalIncomeConfig()
    wb = Workbook()

    # Build the rent roll sheet first to know the annual_total_row index.
    rr_ws = wb.active
    rr_ws.title = SHEET_RENT_ROLL
    annual_total_row = _build_rent_roll_sheet(rr_ws, rows)

    # Insert OER as sheet 0 (leftmost tab).
    oer_ws = wb.create_sheet(SHEET_OER, 0)
    _build_oer_sheet(oer_ws, annual_total_row, _oi)

    # Insert expense detail as sheet 1 (between OER and rent roll).
    exp_ws = wb.create_sheet(SHEET_EXPENSE, 1)
    _build_expense_sheet(exp_ws)

    # Remove the default empty sheet created by Workbook().
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    dest = Path(path)
    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)


# ---------------------------------------------------------------------------
# Sheet builders (private)
# ---------------------------------------------------------------------------

def _build_rent_roll_sheet(ws, rows: list[DirectCapRow]) -> int:
    """Populate 読み取りレントロール.  Returns the annual_total_row (1-based)."""
    headers = [
        "部屋番号", "ステータス",
        "月額賃料", "月額共益費", "水道代収入（月額）", "駐車場収入（月額）", "その他収入（月額）",
        "備考",
    ]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    data_start = 2
    for offset, row in enumerate(rows):
        r = data_start + offset
        note = row.備考
        if note is None and _is_vacant(row.ステータス):
            note = _VACANT_NOTE
        ws.cell(r, _C_ROOM,    row.区画)
        ws.cell(r, _C_STATUS,  row.ステータス or "")
        ws.cell(r, _C_RENT,    row.月額賃料)
        ws.cell(r, _C_CAM,     row.月額共益費)
        ws.cell(r, _C_UTIL,    row.月額水道光熱費)
        ws.cell(r, _C_PARKING, row.月額駐車場)
        ws.cell(r, _C_OTHER,   row.月額その他収入)
        ws.cell(r, _C_NOTES,   note)
        for c in _INCOME_COLS:
            ws.cell(r, c).number_format = _MONEY_FORMAT

    n = len(rows)
    data_end = data_start + n - 1  # last data row (inclusive)
    monthly_row = data_start + n       # row after last data row
    annual_row = monthly_row + 1

    # Monthly total row
    ws.cell(monthly_row, _C_ROOM, "月計")
    ws.cell(monthly_row, _C_NOTES, "月額合計")
    for c in _INCOME_COLS:
        col_letter = get_column_letter(c)
        if n > 0:
            formula = f"=SUM({col_letter}{data_start}:{col_letter}{data_end})"
        else:
            formula = "=0"
        cell = ws.cell(monthly_row, c, formula)
        cell.number_format = _MONEY_FORMAT

    # Annual total row  (monthly * 12, with full border)
    ws.cell(annual_row, _C_ROOM, "年計")
    ws.cell(annual_row, _C_NOTES, "年額合計")
    for c in _INCOME_COLS:
        col_letter = get_column_letter(c)
        monthly_cell_ref = f"{col_letter}{monthly_row}"
        formula = f"={monthly_cell_ref}*12"
        cell = ws.cell(annual_row, c, formula)
        cell.number_format = _MONEY_FORMAT
        cell.border = _FULL_BORDER

    return annual_row


def _build_oer_sheet(
    ws, annual_total_row: int, oi_config: OptionalIncomeConfig | None = None
) -> None:
    """Populate 直接還元法_OER as a self-computing live model.

    Income block (E5:E9):
      E5/E6 (貸室賃料/共益費): always cross-sheet refs to 読み取りレントロール annual row.
      E7-E9 (付帯収入): cross-sheet ref when opt-in, =0 when opt-out.
      Opt-out rows show label + "（算入対象外）" so the user can see what was excluded.

    Input cells (E13:E17): empty with border + light-yellow fill.
    User enters: 空室損失率 / 貸倒損失率 / 経費率 / 資本的支出 / 還元利回り.

    Calculation chain (E20:E24):
      E20 EGI  = E10*(1-N(E13)-N(E14))
      E21 opex = E20*N(E15)          (EGI × expense ratio — main path)
      E22 NOI  = E20-E21
      E23 net  = E22-N(E16)
      E24 value= IFERROR(E23/E17,"") (empty when cap rate is blank)

    Reference rows (E27:E28): 費用詳細版 total for expense-ratio cross-check only.
    These are NOT connected to NOI to avoid double-counting.
    """
    oi = oi_config or OptionalIncomeConfig()
    ann = annual_total_row

    # Title and disclaimer
    ws["A1"] = "直接還元法（収益試算）"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "※ 本値は「収益試算値」であり「収益価格」ではありません。"
    ws["A2"].font = _WARN_FONT

    # Income block header — placed at TOP of income block (D4/E4)
    ws["D4"] = "収入項目"
    ws["E4"] = "年額（円）"
    ws["D4"].font = _BOLD
    ws["E4"].font = _BOLD

    # Income rows E5:E9
    # - optional_key=None (賃料/共益費): always cross-sheet ref
    # - optional_key=str (付帯収入): cross-sheet ref when opt-in, =0 when opt-out
    for row_num, label, income_col, optional_key in _OER_INCOME_ROWS:
        col_letter = get_column_letter(income_col)
        is_optin = (
            optional_key is None
            or (oi.include_in_gpi and optional_key in oi.columns)
        )
        display_label = label if is_optin else f"{label}（算入対象外）"
        ws.cell(row_num, 4, display_label)
        if is_optin:
            formula = f"='{SHEET_RENT_ROLL}'!{col_letter}{ann}"
        else:
            formula = "=0"
        cell = ws.cell(row_num, 5, formula)
        cell.number_format = _MONEY_FORMAT

    # GPI total
    ws["D10"] = "総収入（年額）"
    ws["D10"].font = _BOLD
    ws["E10"] = "=SUM(E5:E9)"
    ws["E10"].number_format = _MONEY_FORMAT
    ws["E10"].font = _BOLD

    # Assumption inputs header
    ws["D12"] = "前提条件（ユーザー入力）"
    ws["D12"].font = _BOLD

    # Input cells E13:E17 — empty, border + light-yellow fill
    input_defs = [
        (13, "空室損失率",           "0.0%"),
        (14, "貸倒損失率",           "0.0%"),
        (15, "経費率（運営費用率）",  "0.0%"),
        (16, "資本的支出（年額）",    _MONEY_FORMAT),
        (17, "還元利回り",            "0.000%"),
    ]
    for row_num, label, fmt in input_defs:
        ws.cell(row_num, 4, label)
        cell = ws.cell(row_num, 5)
        cell.number_format = fmt
        cell.border = _FULL_BORDER
        cell.fill = _INPUT_FILL

    # Calculation results header
    ws["D19"] = "計算結果"
    ws["D19"].font = _BOLD

    ws["D20"] = "有効総収入（EGI）"
    ws["E20"] = "=E10*(1-N(E13)-N(E14))"
    ws["E20"].number_format = _MONEY_FORMAT

    ws["D21"] = "運営費用合計"
    ws["E21"] = "=E20*N(E15)"
    ws["E21"].number_format = _MONEY_FORMAT

    ws["D22"] = "運営純収益（NOI）"
    ws["E22"] = "=E20-E21"
    ws["E22"].number_format = _MONEY_FORMAT

    ws["D23"] = "純収益（還元対象）"
    ws["E23"] = "=E22-N(E16)"
    ws["E23"].number_format = _MONEY_FORMAT

    ws["D24"] = "収益試算値（直接還元法）"
    ws["D24"].font = _BOLD
    ws["E24"] = '=IFERROR(E23/E17,"")'
    ws["E24"].number_format = _MONEY_FORMAT
    ws["E24"].font = _BOLD

    ws["D25"] = "※ 還元利回り（E17）未入力時、収益試算値は空欄になります。"
    ws["D25"].font = _WARN_FONT

    # Reference rows: 費用詳細版 total (参考 — NOT connected to NOI)
    ws["D27"] = "（参考）費用明細合計"
    ws["E27"] = f"='{SHEET_EXPENSE}'!B10"
    ws["E27"].number_format = _MONEY_FORMAT

    ws["D28"] = "（参考）明細ベース経費率"
    ws["E28"] = '=IFERROR(E27/E20,"")'
    ws["E28"].number_format = "0.0%"

    ws["D29"] = (
        "※ 費用明細は出力後にユーザーが入力。"
        "OERのNOIは経費率（E15）で計算され、明細合計（E27）はNOIに連動しません"
        "（経費率の妥当性確認用）。"
    )
    ws["D29"].font = _WARN_FONT


def _build_expense_sheet(ws) -> None:
    """Populate 直接還元法‗費用詳細版 with user-editable structure and SUM row.

    B5:B9 are empty input cells with border + light-yellow fill.
    B10 = SUM(B5:B9) — referenced by OER E27 as a cross-check (not linked to NOI).
    """
    ws["A1"] = "直接還元法 費用詳細版（ユーザー入力）"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = "※ 各費用は実費または想定値を入力してください。"
    ws["A2"].font = _WARN_FONT

    ws["A4"] = "費目"
    ws["B4"] = "年額（円）"
    ws["A4"].font = _BOLD
    ws["B4"].font = _BOLD

    expense_labels = [
        "管理費・管理委託費",
        "修繕費",
        "損害保険料",
        "固定資産税・都市計画税",
        "その他費用",
    ]
    for i, label in enumerate(expense_labels, start=5):
        ws.cell(i, 1, label)
        cell = ws.cell(i, 2)
        cell.number_format = _MONEY_FORMAT
        cell.border = _FULL_BORDER
        cell.fill = _INPUT_FILL

    # SUM row
    ws["A10"] = "費用合計"
    ws["A10"].font = _BOLD
    ws["B10"] = "=SUM(B5:B9)"
    ws["B10"].number_format = _MONEY_FORMAT
    ws["B10"].font = _BOLD

    ws["A12"] = (
        "※ 本シートは出力後にユーザーが入力します。"
        "OERのNOIは経費率（OER!E15）で計算され、本合計は連動しません"
        "（経費率の妥当性確認用）。"
    )
    ws["A12"].font = _WARN_FONT


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _is_vacant(status: str | None) -> bool:
    """Return True if *status* indicates a vacant unit."""
    s = (status or "").strip()
    return any(k in s for k in ("空室", "空き", "募集"))
