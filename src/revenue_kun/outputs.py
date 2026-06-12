"""出力生成: missing_info.md / revenue_analysis.xlsx / extraction_log.json"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import DISCLAIMER, VALUE_LABEL, __version__
from .config import Assumptions
from .missing import MissingItem
from .noi import NOIResult
from .rent_roll import RentRollUnit
from .sensitivity import SensitivityTable
from .valuation import ValuationResult


def _yen(value: float | None) -> str:
    if value is None:
        return "（算定不能）"
    return f"{value:,.0f} 円"


def _pct(value: float | None) -> str:
    if value is None:
        return "（未設定）"
    return f"{value * 100:.3f} %"


# ---------------------------------------------------------------------------
# missing_info.md
# ---------------------------------------------------------------------------
def write_missing_info(path: str | Path, missing: list[MissingItem]) -> None:
    path = Path(path)
    lines: list[str] = []
    lines.append("# 欠損項目一覧 (missing_info.md)")
    lines.append("")
    lines.append(f"> {DISCLAIMER}")
    lines.append("")
    lines.append(
        "本ファイルは、入力データに含まれていなかった項目を列挙したものです。"
        "**これらの項目は推測補完していません。** 必要に応じて一次資料を確認のうえ補完してください。"
    )
    lines.append("")

    if not missing:
        lines.append("欠損項目はありませんでした。")
    else:
        # カテゴリ別にまとめる
        by_cat: dict[str, list[MissingItem]] = {}
        for m in missing:
            by_cat.setdefault(m.category, []).append(m)

        lines.append(f"検出された欠損項目: **{len(missing)} 件**")
        lines.append("")
        for cat, items in by_cat.items():
            lines.append(f"## {cat}（{len(items)} 件）")
            lines.append("")
            lines.append("| 項目 | 出所 | 計算への影響 |")
            lines.append("| --- | --- | --- |")
            for m in items:
                lines.append(f"| {m.field} | {m.location} | {m.impact} |")
            lines.append("")

    path.write_text("\n".join(lines), encoding="utf-8")


# ---------------------------------------------------------------------------
# extraction_log.json
# ---------------------------------------------------------------------------
def _missing_item_dict(m: MissingItem) -> dict:
    return {
        "category": m.category,
        "field": m.field,
        "location": m.location,
        "impact": m.impact,
    }


def write_extraction_log(
    path: str | Path,
    assumptions: Assumptions,
    units: list[RentRollUnit],
    missing: list[MissingItem],
    noi: NOIResult,
    valuation: ValuationResult,
    *,
    input_files: dict[str, str],
    rent_roll_pdf: str | None,
    output_files: dict[str, str],
    executed_at: str,
    extraction_method: str = "dummy",
    phase: str = "Phase 1 (dummy CSV)",
    pdf_extraction: dict | None = None,
) -> None:
    """抽出ログを固定スキーマで JSON 出力する。

    固定スキーマ（必ず出力するトップレベルキー）:
      input_files / rent_roll_pdf / extracted_units_count /
      missing_required_count / missing_optional_count /
      missing_required_items / missing_optional_items /
      gpi / noi / indicated_value / output_files / executed_at
    """
    path = Path(path)

    required_items = [m for m in missing if m.required]
    optional_items = [m for m in missing if not m.required]

    log = {
        # --- メタ情報 ---
        "tool": "収益還元クン (revenue-kun)",
        "version": __version__,
        "phase": phase,
        "disclaimer": DISCLAIMER,
        "extraction_method": extraction_method,  # "dummy"(CSV) / "pdf"
        # --- 固定スキーマ（必須キー） ---
        "input_files": input_files,
        "rent_roll_pdf": rent_roll_pdf,
        "extracted_units_count": len(units),
        "missing_required_count": len(required_items),
        "missing_optional_count": len(optional_items),
        "missing_required_items": [_missing_item_dict(m) for m in required_items],
        "missing_optional_items": [_missing_item_dict(m) for m in optional_items],
        "gpi": noi.gpi,
        "noi": noi.noi,
        "indicated_value": valuation.estimated_value,  # 収益試算値（None=算定不能）
        "output_files": output_files,
        "executed_at": executed_at,
        # --- 付加情報（PDF抽出時のみ） ---
        "pdf_extraction": pdf_extraction or {},
    }

    path.write_text(
        json.dumps(log, ensure_ascii=False, indent=2), encoding="utf-8"
    )


# ---------------------------------------------------------------------------
# extraction_log.json (failure path)
# ---------------------------------------------------------------------------
def write_extraction_failure_log(
    path: str | Path,
    *,
    pdf_name: str,
    failure_reason: str,
    rows_extracted: int = 0,
    pages: int = 0,
    executed_at: str = "",
) -> None:
    """PDF抽出失敗時の最小限の extraction_log.json を出力する。

    成功時の write_extraction_log() と同じファイルに書き込む。
    failure=true と failure_reason を含む簡略スキーマで記録する。
    """
    path = Path(path)
    log = {
        "tool": "収益還元クン (revenue-kun)",
        "version": __version__,
        "disclaimer": DISCLAIMER,
        "extraction_method": "pdf",
        "failure": True,
        "failure_reason": failure_reason,
        "pdf_name": pdf_name,
        "rows_extracted": rows_extracted,
        "pages": pages,
        "executed_at": executed_at,
    }
    path.write_text(json.dumps(log, ensure_ascii=False, indent=2), encoding="utf-8")


# ---------------------------------------------------------------------------
# revenue_analysis.xlsx
# ---------------------------------------------------------------------------
_HEADER_FILL = PatternFill("solid", fgColor="305496")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_TITLE_FONT = Font(bold=True, size=14)
_WARN_FONT = Font(color="C00000")


def _style_header_row(ws, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _autosize(ws, widths: dict[int, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def write_excel(
    path: str | Path,
    assumptions: Assumptions,
    units: list[RentRollUnit],
    noi: NOIResult,
    valuation: ValuationResult,
    sensitivity: SensitivityTable | None,
    missing: list[MissingItem],
) -> None:
    path = Path(path)
    wb = Workbook()

    _sheet_summary(wb, assumptions, noi, valuation)
    _sheet_rent_roll(wb, units)
    _sheet_noi(wb, noi)
    _sheet_sensitivity(wb, sensitivity)
    _sheet_missing(wb, missing)

    # デフォルトの空シートを削除
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(path)


def _sheet_summary(wb, assumptions, noi: NOIResult, valuation: ValuationResult):
    ws = wb.active
    ws.title = "サマリー"

    ws["A1"] = "収益還元クン v0.1 — 収益試算サマリー"
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = DISCLAIMER
    ws["A2"].font = _WARN_FONT
    ws.merge_cells("A2:D2")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = 45

    rows = [
        ("物件名称", assumptions.property_info.get("名称", "")),
        ("所在地", assumptions.property_info.get("所在地", "")),
        ("", ""),
        ("潜在総収入 (GPI)", _yen(noi.gpi)),
        ("空室損失", _yen(-noi.vacancy_loss)),
        ("有効総収入 (EGI)", _yen(noi.egi)),
        ("運営費用合計", _yen(-noi.opex_total)),
        ("運営純収益 (NOI)", _yen(noi.noi)),
        ("資本的支出 (CAPEX)", _yen(-noi.capex)),
        ("純収益（還元対象）", _yen(noi.net_income)),
        ("還元利回り", _pct(valuation.cap_rate)),
        ("", ""),
        (f"{VALUE_LABEL}（直接還元法）", _yen(valuation.estimated_value)),
    ]
    r = 4
    for label, value in rows:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=value)
        if label.startswith(VALUE_LABEL):
            ws.cell(row=r, column=1).font = Font(bold=True, size=12)
            ws.cell(row=r, column=2).font = Font(bold=True, size=12, color="305496")
        r += 1

    ws.cell(row=r + 1, column=1, value="※ 本値は「収益試算値」であり「収益価格」ではありません。")
    ws.cell(row=r + 1, column=1).font = _WARN_FONT

    # 警告
    if noi.warnings:
        r += 3
        ws.cell(row=r, column=1, value="計算上の注意・警告").font = Font(bold=True)
        for w in noi.warnings:
            r += 1
            c = ws.cell(row=r, column=1, value=f"・{w}")
            c.font = _WARN_FONT

    _autosize(ws, {1: 28, 2: 40})


def _sheet_rent_roll(wb, units: list[RentRollUnit]):
    ws = wb.create_sheet("レントロール")
    headers = [
        "区画", "用途", "賃借人", "専有面積(㎡)",
        "月額賃料(円)", "月額共益費(円)", "稼働状況", "契約満了日",
    ]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    for u in units:
        ws.append([
            u.区画, u.用途 or "", u.賃借人 or "",
            u.専有面積_m2, u.月額賃料_円, u.月額共益費_円,
            u.稼働状況 or "", u.契約満了日 or "",
        ])
    # 欠損セルを視覚化
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.value is None:
                cell.value = "（欠損）"
                cell.font = _WARN_FONT

    _autosize(ws, {1: 8, 2: 10, 3: 22, 4: 12, 5: 14, 6: 14, 7: 10, 8: 14})


def _sheet_noi(wb, noi: NOIResult):
    ws = wb.create_sheet("NOI計算")
    ws.append(["項目", "金額(円)", "備考"])
    _style_header_row(ws, 1, 3)

    ws.append(["潜在総収入 (GPI)", noi.gpi, "稼働区画の現行賃料ベース（年額）"])
    ws.append(["空室損失", -noi.vacancy_loss, "GPI × 空室損失率"])
    ws.append(["有効総収入 (EGI)", noi.egi, "GPI − 空室損失"])
    ws.append(["", "", ""])
    ws.append(["【運営費用】", "", ""])
    for name, amount in noi.opex_breakdown.items():
        ws.append([f"  {name}", -amount, ""])
    ws.append(["運営費用合計", -noi.opex_total, "算入できた項目のみ"])
    ws.append(["", "", ""])
    ws.append(["運営純収益 (NOI)", noi.noi, "EGI − 運営費用合計"])
    ws.append(["資本的支出 (CAPEX)", -noi.capex, "控除"])
    ws.append(["純収益（還元対象）", noi.net_income, "NOI − CAPEX"])

    _autosize(ws, {1: 26, 2: 16, 3: 32})


def _sheet_sensitivity(wb, sensitivity: SensitivityTable | None):
    ws = wb.create_sheet("感応度分析")
    if sensitivity is None:
        ws["A1"] = "還元利回りが未設定のため感応度分析を実施できません（補完なし）。"
        ws["A1"].font = _WARN_FONT
        return

    ws["A1"] = f"感応度分析: {VALUE_LABEL}（行=NOI変動率, 列=還元利回り）"
    ws["A1"].font = Font(bold=True)
    ws["A2"] = f"基準: NOI(純収益) {sensitivity.base_net_income:,.0f} 円 / 還元利回り {sensitivity.base_cap_rate*100:.3f}%"

    header_row = 4
    ws.cell(row=header_row, column=1, value="NOI変動率 \\ 還元利回り")
    for j, cap in enumerate(sensitivity.cap_rates, start=2):
        ws.cell(row=header_row, column=j, value=f"{cap*100:.3f}%")
    _style_header_row(ws, header_row, len(sensitivity.cap_rates) + 1)

    for i, rate in enumerate(sensitivity.noi_rates):
        r = header_row + 1 + i
        ws.cell(row=r, column=1, value=f"{rate*100:+.1f}%")
        ws.cell(row=r, column=1).font = _HEADER_FONT
        ws.cell(row=r, column=1).fill = _HEADER_FILL
        for j, val in enumerate(sensitivity.values[i], start=2):
            cell = ws.cell(row=r, column=j, value=val)
            cell.number_format = "#,##0"
            # 基準セル（変動 0×0）を強調
            if sensitivity.noi_rates[i] == 0.0 and abs(sensitivity.cap_rates[j-2] - sensitivity.base_cap_rate) < 1e-9:
                cell.font = Font(bold=True, color="305496")

    widths = {1: 20}
    for j in range(2, len(sensitivity.cap_rates) + 2):
        widths[j] = 16
    _autosize(ws, widths)

    note_r = header_row + len(sensitivity.noi_rates) + 2
    ws.cell(row=note_r, column=1,
            value="※ 各セルは「収益試算値」です。「収益価格」ではありません。").font = _WARN_FONT


def _sheet_missing(wb, missing: list[MissingItem]):
    ws = wb.create_sheet("欠損項目")
    ws.append(["区分", "項目", "出所", "計算への影響"])
    _style_header_row(ws, 1, 4)
    if not missing:
        ws.append(["—", "欠損なし", "—", "—"])
    for m in missing:
        ws.append([m.category, m.field, m.location, m.impact])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    _autosize(ws, {1: 14, 2: 18, 3: 30, 4: 50})
